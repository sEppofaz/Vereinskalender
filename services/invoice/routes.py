import base64
import difflib
import hashlib
import hmac
import json
import os
import re
import tempfile
import threading
import time
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    _OPENPYXL_SUPPORTED = True
except ImportError:
    _OPENPYXL_SUPPORTED = False

import anthropic
import dropbox
from docxtpl import DocxTemplate
from flask import Blueprint, abort, request
from PIL import Image

from shared.kalender_core import MEDIA_TYPES, _HEIC_SUPPORTED, log

invoice_bp = Blueprint("invoice", __name__)

DROPBOX_INVOICE_REFRESH_TOKEN = os.environ["DROPBOX_INVOICE_REFRESH_TOKEN"]
DROPBOX_INVOICE_APP_KEY       = os.environ["DROPBOX_INVOICE_APP_KEY"]
DROPBOX_INVOICE_APP_SECRET    = os.environ["DROPBOX_INVOICE_APP_SECRET"]
CLAUDE_API_KEY                = os.environ["CLAUDE_API_KEY"]
INVOICE_MODEL                 = os.environ.get("CLAUDE_INVOICE_MODEL", "claude-sonnet-4-6")
INVOICE_INPUT_FOLDER          = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/Rechnungen_Input"
INVOICE_OUTPUT_FOLDER         = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/Rechnungen_Entwurf"
INVOICE_DONE_FOLDER           = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/Rechnungen_Erledigt"
INVOICE_ERROR_FOLDER          = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/Rechnungen_Fehler"
INVOICE_CURSOR_FILE           = "/opt/rename-webhook/invoice_cursor.txt"
INVOICE_TEMPLATE              = "/opt/rename-webhook/template.docx"
INVOICE_ADDRESS_FILE          = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/_Adressen.xlsx"
INVOICE_REGISTER_FILE         = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/_Rechnungsregister.xlsx"

_ADDR_HEADERS = ["Name", "Straße", "PLZ", "Ort", "Straße validiert", "PLZ+Ort validiert", "Hinzugefügt"]
_REG_HEADERS  = ["Rechnungsnummer", "Datum", "Anrede", "Nachname", "Vorname",
                 "Produkt", "Netto (€)", "MwSt (€)", "Brutto (€)"]


def get_invoice_dropbox_client() -> dropbox.Dropbox:
    return dropbox.Dropbox(
        oauth2_refresh_token=DROPBOX_INVOICE_REFRESH_TOKEN,
        app_key=DROPBOX_INVOICE_APP_KEY,
        app_secret=DROPBOX_INVOICE_APP_SECRET,
    )


def extract_invoice_data(file_path: str, suffix: str) -> dict:
    """Sendet PDF/Bild an Claude und extrahiert strukturierte Rechnungsdaten."""
    data       = base64.standard_b64encode(Path(file_path).read_bytes()).decode("utf-8")
    media_type = MEDIA_TYPES.get(suffix, "application/pdf")

    system_prompt = (
        "Du bist ein präzises OCR- und Adressvalidierungs-System für handgeschriebene "
        "deutsche Rechnungszettel des Betriebs Josef Kargl, Holzimprägnierwerk, Traich 2, "
        "84101 Obersüßbach (Inhaber: Reinhard Kargl).\n\n"
        "WICHTIG – ABSENDER vs. EMPFÄNGER:\n"
        "Die Zettel sind auf vorgedrucktem Papier geschrieben. Oben links steht gedruckt "
        "'Kargl Reinhard' – das ist der ABSENDER des Betriebs, NICHT der Kunde. "
        "Ignoriere alle gedruckten Texte (Briefkopf, Vordrucke). "
        "Lies ausschließlich die HANDGESCHRIEBENEN Inhalte auf dem Papier. "
        "Der Kundenname und die Kundenadresse stehen immer handgeschrieben.\n\n"
        "ADRESS-VALIDIERUNG – sehr wichtig:\n"
        "Prüfe jeden Adressteil kritisch auf Plausibilität:\n"
        "- Straße: Existiert dieser Straßenname realistisch in Deutschland? "
        "Typische OCR-Fehler bei Handschrift: 'g'→'o', 'n'→'u', 'ei'→'ai', 'rn'→'m' etc. "
        "Beispiele: 'Kirchenwey'→'Kirchenweg', 'Unterolaim'→'Unterglaim', 'Mainbry'→'Mainburg'. "
        "Wenn ein Straßenname ungewöhnliche Buchstabenkombinationen enthält die kein echtes "
        "deutsches Wort ergeben: korrigieren UND address_uncertain=true setzen.\n"
        "- PLZ + Ort: Passt die PLZ zur Region des Ortes? "
        "84xxx = Niederbayern (Landshut, Dingolfing, Mainburg etc.). "
        "Ergolding (84030) liegt bei Landshut – plausibel prüfen.\n"
        "- address_uncertain=true NUR in diesen Fällen: (1) du hast einen Buchstaben korrigiert "
        "weil er als OCR-Fehler erkennbar war, (2) der Straßenname oder Ort ist unleserlich, "
        "(3) PLZ und Ort passen geografisch nicht zusammen. "
        "address_uncertain=false wenn die Adresse klar lesbar und geografisch plausibel ist – "
        "auch wenn die Handschrift etwas schwer lesbar war aber eindeutig entzifferbar.\n\n"
        "WICHTIG: Berechne KEINE Summen selbst. Lies nur die Rohdaten vom Zettel ab.\n\n"
        "Gib AUSSCHLIESSLICH ein valides JSON-Objekt zurück – keinen weiteren Text, "
        "keine Erklärungen, keine Markdown-Backticks."
    )

    user_prompt = (
        "Extrahiere alle handgeschriebenen Daten aus diesem Rechnungszettel "
        "und gib folgendes JSON zurück:\n"
        "{\n"
        "  \"anrede\": \"Firma | Herr | Frau\",\n"
        "  \"name\": \"vollständiger Name des KUNDEN (handgeschrieben, nicht der gedruckte Briefkopf)\",\n"
        "  \"strasse_nr\": \"Straße und Hausnummer des Kunden\",\n"
        "  \"plz\": \"Postleitzahl des Kunden\",\n"
        "  \"ort\": \"Ort des Kunden\",\n"
        "  \"address_uncertain\": false,\n"
        "  \"beschreibungstext\": \"Beschreibungstext für die Rechnung. "
        "Format: 'Wir [VERB] in Ihrem Auftrag [nachstehende/nachstehendes] [MATERIAL] [in KW X / am DD.MM.YYYY]:' – "
        "VERB: exakt vom Zettel lesen (z.B. 'imprägnierten', 'schälten', 'hobelten'). "
        "MATERIAL: exakter Materialname vom Zettel (z.B. 'Schnittholz', 'Bretter', 'Dachlatten 30x50 mm', 'Rundholz'). "
        "Grammatik beachten: 'nachstehendes' bei Neutrum (Rundholz, Schnittholz), 'nachstehende' bei Plural (Bretter, Dachlatten). "
        "Niemals Verb oder Material erfinden. Wenn kein Datum/KW lesbar: weglassen.\",\n"
        "  \"positionen\": [\n"
        "    {\"menge\": 1.66, \"einheit\": \"cbm\", \"einzelpreis\": 110.00, \"positions_beschreibung\": \"\"}\n"
        "  ],\n"
        "  \"netto_auf_zettel\": 182.60,\n"
        "  \"mwst_auf_zettel\": 34.69,\n"
        "  \"brutto_auf_zettel\": 217.29,\n"
        "  \"hinweis\": \"Betrag bereits bar bezahlt.\"\n"
        "}\n\n"
        "Hinweise:\n"
        "- name: das ist IMMER der handgeschriebene Kundenname – niemals 'Kargl' oder 'Reinhard'\n"
        "- anrede: 'Firma' wenn GbR, GmbH, AG o.ä., sonst 'Herr' oder 'Frau'\n"
        "- positionen: ein Eintrag pro Zeile auf dem Zettel, wenn Menge UND Einzelpreis angegeben sind. "
        "Einheit übernehmen wie auf dem Zettel (z.B. 'cbm', 'St.', 'lfm', 'fm'). "
        "positions_beschreibung: zusätzliche Spezifikationen der Position (z.B. '9m, 20-22 cm Zopf', "
        "'4m lang', '30x50 mm') – leer lassen wenn keine vorhanden. "
        "Wenn KEIN Einzelpreis angegeben ist (nur ein Gesamtbetrag): positionen = []\n"
        "- netto_auf_zettel / mwst_auf_zettel / brutto_auf_zettel: alle auf dem Zettel notierten "
        "Beträge eintragen (null wenn nicht lesbar/vorhanden)\n"
        "- beschreibungstext: Verb UND Material EXAKT aus der Handschrift lesen – "
        "niemals Verb oder Material erfinden; Grammatik beachten (nachstehendes/nachstehende)\n"
        "- hinweis: NUR explizite Zahlungshinweise vom Zettel (z.B. 'Betrag bereits bar bezahlt.', "
        "'Bereits überwiesen.') – Telefonnummern, E-Mail-Adressen, betriebliche Notizen und "
        "sonstige Vermerke gehören NICHT in den Hinweis. "
        "Leer lassen wenn kein Zahlungshinweis vorhanden.\n"
        "- beschreibungstext: Den Leistungstext formulieren. "
        "Wenn Verb und Material klar erkennbar sind: "
        "'Wir [VERB] in Ihrem Auftrag [nachstehende/nachstehendes] [MATERIAL] [Datum/KW]:'. "
        "Wenn kein klares Verb/Material vorhanden, sondern nur Notizen oder Referenzen "
        "(z.B. 'OKK als Ersatzteilspender', 'OKK TTH 4 Bagger'): "
        "diese Notizen direkt als Beschreibungstext übernehmen, exakt wie geschrieben. "
        "Telefonnummern und E-Mail-Adressen gehören NICHT in den Beschreibungstext.\n"
        "- address_uncertain=true NUR wenn ein Buchstabe korrigiert wurde oder der Ort "
        "nicht eindeutig lesbar war – NICHT bei klar lesbaren Adressen wie 'Industriestr. 2, "
        "94330 Salching'"
    )

    if suffix == ".pdf":
        file_content = {"type": "document", "source": {"type": "base64", "media_type": media_type, "data": data}}
    else:
        file_content = {"type": "image",    "source": {"type": "base64", "media_type": media_type, "data": data}}

    client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
    raw    = None
    for attempt in range(1, 4):
        try:
            message = client.messages.create(
                model=INVOICE_MODEL,
                max_tokens=1024,
                system=system_prompt,
                messages=[{"role": "user", "content": [file_content, {"type": "text", "text": user_prompt}]}],
            )
            raw = message.content[0].text.strip()
            break
        except anthropic.APIStatusError as e:
            if e.status_code == 529 and attempt < 3:
                wait = attempt * 15
                log(f"⏳  API überlastet (Versuch {attempt}/3) – warte {wait}s ...")
                time.sleep(wait)
            else:
                raise

    if "```" in raw:
        match = re.search(r"```(?:json)?\s*([\s\S]+?)```", raw)
        raw = match.group(1).strip() if match else raw

    if not raw:
        raise ValueError("Claude hat eine leere Antwort geliefert")

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        log(f"⚠️  Claude-Antwort (kein JSON): {raw[:300]}")
        raise


def calculate_and_validate(data: dict) -> dict:
    positionen    = [p for p in (data.get("positionen") or [])
                     if (p.get("menge") or p.get("menge_cbm")) and p.get("einzelpreis")]
    netto_zettel  = data.get("netto_auf_zettel")
    mwst_zettel   = data.get("mwst_auf_zettel")
    brutto_zettel = data.get("brutto_auf_zettel")

    if positionen:
        netto  = round(sum((p.get("menge") or p.get("menge_cbm")) * p["einzelpreis"] for p in positionen), 2)
        mwst   = round(netto * 0.19, 2)
        brutto = round(netto + mwst, 2)

        netto_ok  = netto_zettel  is None or abs(netto  - netto_zettel)  < 0.02
        brutto_ok = brutto_zettel is None or abs(brutto - brutto_zettel) < 0.02

        if not netto_ok:
            log(f"⚠️  Netto-Abweichung: berechnet {netto:.2f} € vs. Zettel {netto_zettel:.2f} €")
        if not brutto_ok:
            log(f"⚠️  Brutto-Abweichung: berechnet {brutto:.2f} € vs. Zettel {brutto_zettel:.2f} €")
    else:
        log("ℹ️  Pauschalbetrag-Modus: keine cbm×Preis-Positionen, Zettelwerte werden verwendet")
        brutto      = brutto_zettel or 0.0
        netto       = netto_zettel  or round(brutto / 1.19, 2)
        mwst        = mwst_zettel   or round(netto * 0.19, 2)
        brutto_calc = round(netto + mwst, 2)
        netto_ok    = True
        brutto_ok   = abs(brutto_calc - brutto) < 0.02
        if not brutto_ok:
            log(f"⚠️  Pauschalbetrag-Abweichung: {netto:.2f} + {mwst:.2f} = {brutto_calc:.2f} "
                f"≠ Zettel {brutto:.2f} €")

    return {"netto": netto, "mwst": mwst, "brutto": brutto,
            "netto_ok": netto_ok, "brutto_ok": brutto_ok, "pauschal": not bool(positionen)}


def fmt_eur(value: float) -> str:
    formatted = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{formatted} €"


def fmt_cbm(value: float) -> str:
    return f"{value:.2f}".replace(".", ",") + " cbm"


def build_docx(data: dict, calc: dict, rechnungsnummer: str = "") -> str:
    tpl        = DocxTemplate(INVOICE_TEMPLATE)
    positionen = [p for p in (data.get("positionen") or [])
                  if (p.get("menge") or p.get("menge_cbm")) and p.get("einzelpreis")]

    pos_context = {}
    if positionen:
        for i in range(1, 7):
            if i <= len(positionen):
                p            = positionen[i - 1]
                menge        = p.get("menge") or p.get("menge_cbm")
                einheit      = p.get("einheit") or "cbm"
                pos_zusatz   = p.get("positions_beschreibung", "") or ""
                zeilen_netto = round(menge * p["einzelpreis"], 2)
                if einheit == "cbm":
                    pos_str = fmt_cbm(menge)
                elif menge == int(menge):
                    pos_str = f"{int(menge)} {einheit}"
                else:
                    pos_str = f"{menge:.2f}".replace(".", ",") + f" {einheit}"
                if pos_zusatz:
                    pos_str += f", {pos_zusatz}"
                pos_context[f"position{i}"]    = pos_str
                pos_context[f"einzelpreis{i}"] = fmt_eur(p["einzelpreis"])
                pos_context[f"gesamtpreis{i}"] = fmt_eur(zeilen_netto)
            else:
                pos_context[f"position{i}"]    = ""
                pos_context[f"einzelpreis{i}"] = ""
                pos_context[f"gesamtpreis{i}"] = ""
    else:
        pos_context["position1"]    = ""
        pos_context["einzelpreis1"] = ""
        pos_context["gesamtpreis1"] = fmt_eur(calc["netto"])
        for i in range(2, 7):
            pos_context[f"position{i}"]    = ""
            pos_context[f"einzelpreis{i}"] = ""
            pos_context[f"gesamtpreis{i}"] = ""

    context = {
        "rechnungsnummer":   rechnungsnummer,
        "anrede":            data.get("anrede", ""),
        "name":              data.get("name", ""),
        "strasse_nr":        data.get("strasse_nr", ""),
        "plz":               data.get("plz", ""),
        "ort":               data.get("ort", ""),
        "datum":             datetime.now().strftime("%d.%m.%Y"),
        "beschreibungstext": data.get("beschreibungstext", ""),
        "hinweis":           data.get("hinweis") or "",
        "netto":             fmt_eur(calc["netto"]),
        "mwst":              fmt_eur(calc["mwst"]),
        "brutto":            fmt_eur(calc["brutto"]),
        **pos_context,
    }

    tpl.render(context)
    tmp_path = tempfile.mktemp(suffix=".docx")
    tpl.save(tmp_path)
    return tmp_path


def _download_excel(dbx: dropbox.Dropbox, dropbox_path: str):
    _, res = dbx.files_download(dropbox_path)
    tmp = tempfile.mktemp(suffix=".xlsx")
    Path(tmp).write_bytes(res.content)
    wb = openpyxl.load_workbook(tmp)
    Path(tmp).unlink(missing_ok=True)
    return wb


def _upload_excel(dbx: dropbox.Dropbox, wb, dropbox_path: str) -> None:
    tmp = tempfile.mktemp(suffix=".xlsx")
    wb.save(tmp)
    with open(tmp, "rb") as f:
        dbx.files_upload(f.read(), dropbox_path, mode=dropbox.files.WriteMode.overwrite, mute=True)
    Path(tmp).unlink(missing_ok=True)


def _ensure_address_excel(dbx: dropbox.Dropbox):
    try:
        return _download_excel(dbx, INVOICE_ADDRESS_FILE)
    except dropbox.exceptions.ApiError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Adressen"
        ws.append(_ADDR_HEADERS)
        header_fill = PatternFill("solid", fgColor="4472C4")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 15
        _upload_excel(dbx, wb, INVOICE_ADDRESS_FILE)
        log("📋  Adressen.xlsx neu angelegt")
        return wb


def find_in_address_excel(dbx: dropbox.Dropbox, name: str) -> dict | None:
    try:
        wb         = _ensure_address_excel(dbx)
        ws         = wb.active
        name_lower = name.strip().lower()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip().lower() == name_lower:
                return {"name": row[0], "strasse_nr": row[1], "plz": str(row[2]), "ort": row[3]}
    except Exception as e:
        log(f"⚠️  Adressliste lesen: {e}")
    return None


def add_to_address_excel(dbx: dropbox.Dropbox, name: str, strasse: str,
                         plz: str, ort: str, street_ok: bool, location_ok: bool,
                         street_corrected: bool = False) -> None:
    try:
        wb          = _ensure_address_excel(dbx)
        ws          = wb.active
        street_val  = "Korrigiert" if street_corrected else ("Ja" if street_ok else "Nein")
        ws.append([name, strasse, plz, ort,
                   street_val, "Ja" if location_ok else "Nein",
                   datetime.now().strftime("%Y-%m-%d")])
        _upload_excel(dbx, wb, INVOICE_ADDRESS_FILE)
        log(f"📋  Adresse hinzugefügt: {name}")
    except Exception as e:
        log(f"⚠️  Adresse speichern: {e}")


def _format_invoice_nr(seq: int, year: int) -> str:
    return f"{seq:03d}00{year:02d}"


def _parse_invoice_nr(nr_str: str) -> tuple[int, int]:
    nr = str(nr_str).strip()
    if len(nr) == 7:
        return int(nr[:3]), int(nr[5:7])
    return 0, 0


def get_next_invoice_number(dbx: dropbox.Dropbox) -> str:
    current_year = int(datetime.now().strftime("%y"))
    try:
        wb      = _download_excel(dbx, INVOICE_REGISTER_FILE)
        ws      = wb.active
        last_nr = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                last_nr = str(row[0])
        if last_nr:
            seq, year = _parse_invoice_nr(last_nr)
            if year == current_year:
                return _format_invoice_nr(seq + 1, current_year)
            return _format_invoice_nr(1, current_year)
    except dropbox.exceptions.ApiError:
        pass
    except Exception as e:
        log(f"⚠️  Rechnungsnummer lesen: {e}")
    return _format_invoice_nr(17, current_year)


def _ensure_register_excel(dbx: dropbox.Dropbox):
    try:
        return _download_excel(dbx, INVOICE_REGISTER_FILE)
    except dropbox.exceptions.ApiError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rechnungen"
        ws.append(_REG_HEADERS)
        header_fill = PatternFill("solid", fgColor="375623")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for col, width in zip("ABCDEFGHI", [16, 12, 10, 22, 18, 45, 12, 12, 12]):
            ws.column_dimensions[col].width = width
        log("📊  Rechnungsregister.xlsx neu angelegt")
        return wb


def save_to_invoice_register(dbx: dropbox.Dropbox, rechnungsnummer: str,
                              data: dict, calc: dict) -> None:
    try:
        wb      = _ensure_register_excel(dbx)
        ws      = wb.active
        anrede  = data.get("anrede", "")
        name    = data.get("name", "")
        if anrede == "Firma":
            nachname, vorname = name, ""
        else:
            parts    = name.strip().rsplit(" ", 1)
            nachname = parts[1] if len(parts) == 2 else name
            vorname  = parts[0] if len(parts) == 2 else ""
        beschreibung = (data.get("beschreibungstext") or "")[:60]
        ws.append([
            rechnungsnummer, datetime.now().strftime("%d.%m.%Y"),
            anrede, nachname, vorname, beschreibung,
            round(calc["netto"], 2), round(calc["mwst"], 2), round(calc["brutto"], 2),
        ])
        _upload_excel(dbx, wb, INVOICE_REGISTER_FILE)
        log(f"📊  Register: {rechnungsnummer} – {name}")
    except Exception as e:
        log(f"⚠️  Rechnungsregister speichern: {e}")


def _nominatim_get(url: str) -> list:
    time.sleep(1)
    req = urllib.request.Request(url, headers={"User-Agent": "KarglRechnungsService/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            return json.loads(r.read())
    except Exception as e:
        log(f"⚠️  Nominatim: {e}")
        return []


def validate_and_correct_address(strasse_nr: str, plz: str, ort: str) -> dict:
    query   = f"{strasse_nr}, {plz} {ort}, Deutschland"
    url     = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
        {"q": query, "format": "json", "limit": "1", "countrycodes": "de", "addressdetails": "1"})
    results = _nominatim_get(url)

    out = {"street_ok": False, "strasse_nr": strasse_nr, "plz": plz, "ort": ort, "corrected": False}
    if not results:
        return out

    out["street_ok"] = True
    addr = results[0].get("address", {})

    osm_road = addr.get("road", "")
    osm_hnr  = addr.get("house_number", "")
    osm_plz  = addr.get("postcode", "")

    ort_candidates = [v for k, v in addr.items()
                      if k in ("city", "town", "municipality", "village", "hamlet", "suburb", "district") and v]
    if ort_candidates:
        best_ort   = max(ort_candidates, key=lambda c: difflib.SequenceMatcher(None, ort.lower(), c.lower()).ratio())
        best_ratio = difflib.SequenceMatcher(None, ort.lower(), best_ort.lower()).ratio()
        osm_ort    = best_ort if best_ratio >= 0.6 else ""
    else:
        osm_ort = ""

    m         = re.match(r"^(.*?)\s+(\d+\w*)$", strasse_nr.strip())
    input_hnr = m.group(2) if m else ""

    hnr         = osm_hnr or input_hnr
    new_strasse = f"{osm_road} {hnr}".strip() if osm_road else strasse_nr
    new_plz     = osm_plz or plz
    new_ort     = osm_ort or ort

    if (new_strasse.lower() != strasse_nr.strip().lower()
            or new_plz != plz
            or new_ort.lower() != ort.lower()):
        out["corrected"] = True

    out["strasse_nr"] = new_strasse
    out["plz"]        = new_plz
    out["ort"]        = new_ort
    return out


def validate_location_nominatim(plz: str, ort: str) -> bool:
    query = f"{plz} {ort}, Deutschland"
    url   = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
        {"q": query, "format": "json", "limit": "1", "countrycodes": "de"})
    return bool(_nominatim_get(url))


def lookup_company_nominatim(firma: str, strasse_nr: str, plz: str, ort: str) -> str | None:
    for query in [f"{firma}, {strasse_nr}, {plz} {ort}", f"{firma}, {ort}, Deutschland"]:
        url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
            {"q": query, "format": "json", "limit": "3", "countrycodes": "de"})
        results = _nominatim_get(url)
        if results:
            osm_name = results[0].get("name", "").strip()
            if osm_name:
                ratio = difflib.SequenceMatcher(None, firma.lower(), osm_name.lower()).ratio()
                if ratio >= 0.5:
                    return osm_name
    return None


def enrich_invoice_address(dbx: dropbox.Dropbox, data: dict) -> dict:
    name  = data.get("name", "").strip()
    known = find_in_address_excel(dbx, name)
    if known:
        data["strasse_nr"]        = known["strasse_nr"]
        data["plz"]               = known["plz"]
        data["ort"]               = known["ort"]
        data["address_uncertain"] = False
        log(f"📋  Adresse aus Liste übernommen: {name}")
        return data

    strasse     = data.get("strasse_nr", "")
    plz         = data.get("plz", "")
    ort         = data.get("ort", "")
    location_ok = validate_location_nominatim(plz, ort)
    addr        = validate_and_correct_address(strasse, plz, ort)

    if not location_ok:
        data["address_uncertain"] = True
        log(f"⚠️  PLZ+Ort nicht verifiziert: {plz} {ort}")
    elif not addr["street_ok"]:
        data["address_uncertain"] = True
        log(f"⚠️  Straße nicht verifiziert: {strasse}, {plz} {ort}")
    else:
        if addr["corrected"]:
            log(f"🔧  Adresse korrigiert: '{strasse}, {plz} {ort}' → "
                f"'{addr['strasse_nr']}, {addr['plz']} {addr['ort']}'")
            data["address_uncertain"] = True
        else:
            log(f"✅  Adresse bestätigt: {addr['strasse_nr']}, {addr['plz']} {addr['ort']}")
        data["strasse_nr"] = addr["strasse_nr"]
        data["plz"]        = addr["plz"]
        data["ort"]        = addr["ort"]

    if data.get("anrede") == "Firma":
        osm_name = lookup_company_nominatim(
            name, data.get("strasse_nr", strasse), data.get("plz", plz), data.get("ort", ort))
        if osm_name and osm_name != name:
            log(f"🏢  Firmenname korrigiert: '{name}' → '{osm_name}'")
            data["name"] = osm_name
            name = osm_name
        elif not osm_name:
            data["address_uncertain"] = True
            log(f"⚠️  Firma nicht in OSM gefunden: {name}")

    add_to_address_excel(dbx, name,
                         data.get("strasse_nr", strasse), data.get("plz", plz), data.get("ort", ort),
                         addr["street_ok"], location_ok, addr["corrected"])
    return data


def _move_to_error(dbx: dropbox.Dropbox, dropbox_path: str, filename: str) -> None:
    try:
        error_path = f"{INVOICE_ERROR_FOLDER}/{filename}"
        dbx.files_move_v2(dropbox_path, error_path, autorename=True)
        log(f"⚠️  Nach Fehler verschoben: {INVOICE_ERROR_FOLDER}/{filename}")
    except Exception as e:
        log(f"❌  Konnte Datei nicht nach Fehler-Ordner verschieben: {e}")


def process_invoice(dbx: dropbox.Dropbox, dropbox_path: str) -> None:
    filename = Path(dropbox_path).name
    suffix   = Path(dropbox_path).suffix.lower()

    if suffix not in {".pdf", ".jpg", ".jpeg", ".png", ".heic", ".heif"}:
        log(f"ℹ️  Invoice: überspringe {filename} (kein unterstütztes Format)")
        return
    if suffix in {".heic", ".heif"} and not _HEIC_SUPPORTED:
        log(f"⚠️  Invoice: HEIC-Datei {filename} empfangen, aber pillow-heif nicht installiert – überspringe")
        return

    log(f"🧾  Starte Rechnungsverarbeitung: {filename}")

    tmp_file      = None
    tmp_docx      = None
    tmp_converted = None

    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as f:
            tmp_file = f.name
        dbx.files_download_to_file(tmp_file, dropbox_path)
        log(f"⬇️  Heruntergeladen: {filename}")

        process_path   = tmp_file
        process_suffix = suffix
        if suffix in {".heic", ".heif"}:
            tmp_converted = tempfile.mktemp(suffix=".jpg")
            with Image.open(tmp_file) as img:
                img.convert("RGB").save(tmp_converted, format="JPEG", quality=92)
            process_path   = tmp_converted
            process_suffix = ".jpg"
            log(f"🔄  HEIC → JPEG konvertiert: {filename}")

        invoice_data    = extract_invoice_data(process_path, process_suffix)
        n_pos           = len(invoice_data.get("positionen") or [])
        log(f"🤖  Extrahiert: {invoice_data.get('name', '?')} – {n_pos} Position(en)")

        rechnungsnummer = get_next_invoice_number(dbx)
        log(f"🔢  Rechnungsnummer: {rechnungsnummer}")

        invoice_data = enrich_invoice_address(dbx, invoice_data)
        calc         = calculate_and_validate(invoice_data)
        tmp_docx     = build_docx(invoice_data, calc, rechnungsnummer)

        clean_name   = re.sub(r"[^\w\-]", "_", invoice_data.get("name", "Unbekannt"))
        datum_str    = datetime.now().strftime("%Y-%m-%d")
        brutto_str   = f"{calc['brutto']:.2f}"
        needs_prufen = (invoice_data.get("address_uncertain")
                        or not calc["netto_ok"] or not calc["brutto_ok"])
        nr_prefix    = f"_prüfen_{rechnungsnummer}" if needs_prufen else rechnungsnummer
        out_name     = f"{nr_prefix}_Rechnung_{clean_name}_{datum_str}_{brutto_str}€.docx"
        out_path     = f"{INVOICE_OUTPUT_FOLDER}/{out_name}"

        with open(tmp_docx, "rb") as f:
            dbx.files_upload(f.read(), out_path, mode=dropbox.files.WriteMode.overwrite, mute=True)
        log(f"⬆️  Hochgeladen: {out_path}")

        save_to_invoice_register(dbx, rechnungsnummer, invoice_data, calc)

        done_name = f"{rechnungsnummer}_Eingang_{clean_name}_{datum_str}_{brutto_str}€{suffix}"
        done_path = f"{INVOICE_DONE_FOLDER}/{done_name}"
        dbx.files_move_v2(dropbox_path, done_path, autorename=True)
        log(f"📁  Archiviert: {done_path}")

    except json.JSONDecodeError as e:
        log(f"❌  JSON-Fehler bei {filename}: {e}")
        _move_to_error(dbx, dropbox_path, filename)
    except Exception as e:
        log(f"❌  Fehler bei Rechnungsverarbeitung {filename}: {e}")
        _move_to_error(dbx, dropbox_path, filename)
    finally:
        if tmp_file:      Path(tmp_file).unlink(missing_ok=True)
        if tmp_converted: Path(tmp_converted).unlink(missing_ok=True)
        if tmp_docx:      Path(tmp_docx).unlink(missing_ok=True)


def process_invoice_changes(dbx: dropbox.Dropbox) -> None:
    from dropbox.files import DeletedMetadata, FolderMetadata
    cursor = get_invoice_cursor(dbx)
    result = dbx.files_list_folder_continue(cursor)

    while True:
        for entry in result.entries:
            if isinstance(entry, (DeletedMetadata, FolderMetadata)):
                continue
            if str(Path(entry.path_lower).parent) != INVOICE_INPUT_FOLDER.lower():
                continue
            try:
                process_invoice(dbx, entry.path_display)
            except Exception as e:
                log(f"❌  Invoice-Fehler bei {entry.name}: {e}")

        save_invoice_cursor(result.cursor)
        if not result.has_more:
            break
        result = dbx.files_list_folder_continue(result.cursor)


def get_invoice_cursor(dbx: dropbox.Dropbox) -> str:
    if Path(INVOICE_CURSOR_FILE).exists():
        return Path(INVOICE_CURSOR_FILE).read_text().strip()
    result = dbx.files_list_folder(INVOICE_INPUT_FOLDER, recursive=False)
    cursor = result.cursor
    while result.has_more:
        result = dbx.files_list_folder_continue(cursor)
        cursor = result.cursor
    Path(INVOICE_CURSOR_FILE).write_text(cursor)
    log("🔖  Invoice-Cursor gespeichert")
    return cursor


def save_invoice_cursor(cursor: str) -> None:
    Path(INVOICE_CURSOR_FILE).write_text(cursor)


@invoice_bp.route("/webhook-invoice", methods=["GET"])
def verify_invoice():
    challenge = request.args.get("challenge")
    if challenge:
        return challenge, 200, {"Content-Type": "text/plain"}
    abort(400)


@invoice_bp.route("/webhook-invoice", methods=["POST"])
def webhook_invoice():
    signature = request.headers.get("X-Dropbox-Signature", "")
    expected  = hmac.new(
        DROPBOX_INVOICE_APP_SECRET.encode(),
        request.data,
        hashlib.sha256,
    ).hexdigest()
    if not hmac.compare_digest(signature, expected):
        log("⚠️  Ungültige Invoice-Webhook-Signatur")
        abort(403)

    dbx = get_invoice_dropbox_client()
    threading.Thread(target=process_invoice_changes, args=(dbx,), daemon=True).start()
    return "", 200
