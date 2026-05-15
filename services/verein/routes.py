import html
import io
import json
import os
import re
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path

from flask import Blueprint, make_response, redirect, request

from shared.kalender_store import KalenderStore
from shared.kalender_core import (
    VEREINSTERMINE_FILE, _HEIC_SUPPORTED, _do_save_import,
    import_pdf_bytes, lookup_plz, parse_excel_bytes,
)
from shared.vk_db import (
    db_conn, get_session_user, log_audit,
    get_upload_count, increment_upload_quota,
)
from services.auth.routes import _CSS, _page, _session_token, require_verein_login

verein_bp = Blueprint("verein", __name__)

_BACK_DASH = '<a class="btn btn-sec" href="/verein/dashboard" style="margin-top:.75rem">← Zurück</a>'
_UPLOAD_LIMIT = 3
RUBRIKEN = ["Verein", "Pfarrei", "Kunst und Kultur", "Sonstiges"]


def _quota_remaining(verein_id: int) -> int:
    return max(0, _UPLOAD_LIMIT - get_upload_count(verein_id, date.today().isoformat()))


def _load_data() -> dict:
    return json.loads(VEREINSTERMINE_FILE.read_text())


def _get_verein_termine(verein_key: str) -> list:
    data = _load_data()
    alle = data.get(verein_key, [])
    return [t for t in alle if not t.get("deleted")]


# ── Dashboard ────────────────────────────────────────────────────────────────

@verein_bp.route("/verein/dashboard")
@require_verein_login
def dashboard(user):
    verein_key = user["verein_key"]
    verein_name = user["verein_name"]
    termine = _get_verein_termine(verein_key)
    heute = date.today().isoformat()

    rows = ""
    for t in sorted(termine, key=lambda x: x.get("datum", "")):
        past = t.get("datum", "") < heute
        style = "opacity:.5" if past else ""
        edit_btn = ""
        if user["role"] == "admin":
            edit_btn = f'<a href="/verein/termine/{t["id"]}" style="margin-left:.5rem;color:#0a84ff;text-decoration:none">✏️</a>'
        rows += f"""<div class="card" style="{style}">
  <div style="display:flex;justify-content:space-between;align-items:start">
    <div>
      <div style="font-weight:600">{t.get('bezeichnung','')}</div>
      <div style="color:#aeaeb2;font-size:.85rem">{t.get('datum','')} {t.get('uhrzeit','')}</div>
      <div style="color:#aeaeb2;font-size:.85rem">{t.get('ort','')}</div>
    </div>
    <div>{edit_btn}</div>
  </div>
</div>"""

    if not rows:
        rows = '<p style="color:#aeaeb2">Noch keine Termine eingetragen.</p>'

    neu_btn = ""
    upload_btn = ""
    mitglieder_link = ""
    profil_link = ""
    if user["role"] == "admin":
        neu_btn = '<a class="btn" href="/verein/termine/neu">+ Neuer Termin</a>'
        remaining = _quota_remaining(user["verein_id"])
        used = _UPLOAD_LIMIT - remaining
        upload_btn = (
            f'<a class="btn btn-sec" href="/verein/upload" style="margin-top:.5rem">'
            f'📤 Terminplan hochladen ({used}/{_UPLOAD_LIMIT} heute)</a>'
        )
        mitglieder_link = '<a class="btn btn-sec" href="/verein/mitglieder" style="margin-top:.5rem">👥 Mitglieder</a>'
        profil_link = '<a class="btn btn-sec" href="/verein/profil" style="margin-top:.5rem">⚙️ Vereinsprofil</a>'

    upload_ok = request.args.get("upload_ok", "")
    upload_banner = ""
    if upload_ok and upload_ok.isdigit():
        upload_banner = f'<p class="ok">✅ {upload_ok} Termine erfolgreich importiert.</p>'

    hilfe_block = ""
    if user["role"] == "admin":
        hilfe_block = """
<details style="margin-top:1rem;border:1px solid #3a3a3c;border-radius:.625rem;overflow:hidden">
  <summary style="padding:.75rem 1rem;cursor:pointer;background:#2c2c2e;color:#f2f2f7;font-size:.9rem;font-weight:600;list-style:none;display:flex;justify-content:space-between;align-items:center">
    ❓ Hilfe &amp; FAQ <span style="color:#aeaeb2;font-weight:400;font-size:.8rem">▾</span>
  </summary>
  <div style="padding:1rem;display:flex;flex-direction:column;gap:.85rem;background:#1c1c1e">

    <div>
      <div style="font-weight:600;font-size:.9rem;margin-bottom:.25rem">📄 PDF oder Foto hochladen</div>
      <div style="color:#aeaeb2;font-size:.85rem">Claude KI liest das Dokument und extrahiert Termine automatisch. Funktioniert mit Jahresprogrammen, Pfarrbriefen und Fotos von Plakaten (JPG, PNG, HEIC). Dauer: ca. 15–60 Sek.</div>
    </div>

    <div>
      <div style="font-weight:600;font-size:.9rem;margin-bottom:.25rem">📊 Excel-Vorlage</div>
      <div style="color:#aeaeb2;font-size:.85rem">Vorlage herunterladen, ausfüllen und hochladen – kein KI-Call, sofortige Verarbeitung.<br>
      Datumsformat: <code style="background:#3a3a3c;padding:0 4px;border-radius:3px">TT.MM.JJJJ</code> oder <code style="background:#3a3a3c;padding:0 4px;border-radius:3px">JJJJ-MM-TT</code> – kein Text wie „ca." oder Leerzeichen in der Datumsspalte.</div>
    </div>

    <div>
      <div style="font-weight:600;font-size:.9rem;margin-bottom:.25rem">⏱ Tageslimit</div>
      <div style="color:#aeaeb2;font-size:.85rem">3 Uploads pro Tag – Zurücksetzung um Mitternacht. Das Limit gilt pro Verein.</div>
    </div>

    <div>
      <div style="font-weight:600;font-size:.9rem;margin-bottom:.25rem">🔄 Upload fehlgeschlagen?</div>
      <div style="color:#aeaeb2;font-size:.85rem">
        <b style="color:#f2f2f7">PDF:</b> Seite als Foto abfotografieren und als JPG hochladen.<br>
        <b style="color:#f2f2f7">Excel:</b> Datumsspalte prüfen – nur reines Datum, kein zusätzlicher Text.<br>
        <b style="color:#f2f2f7">Allgemein:</b> Datei erneut hochladen oder per Mail an
        <a href="mailto:Vereinskalender@icloud.com" style="color:#0a84ff">Vereinskalender@icloud.com</a> schicken – wir importieren manuell.
      </div>
    </div>

  </div>
</details>"""

    body = f"""
{upload_banner}<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:1rem">
  <div>
    <div style="font-weight:600">{verein_name}</div>
    <div style="color:#aeaeb2;font-size:.85rem">{user['email']} · {user['role']}</div>
  </div>
  <form method="post" action="/verein/logout">
    <button class="btn btn-sec" style="width:auto;padding:.5rem .875rem;font-size:.85rem">Logout</button>
  </form>
</div>
{neu_btn}
{upload_btn}
<h2 style="font-size:1rem;margin:1rem 0 .5rem">Termine ({len(termine)})</h2>
{rows}
{mitglieder_link}
{profil_link}
{hilfe_block}
<hr>
<a class="btn btn-sec" href="/verein/passwort" style="margin-top:.5rem">🔑 Passwort ändern</a>
<a class="btn btn-sec" href="/" style="margin-top:.5rem">← Zurück zum Kalender</a>
<p class="hint" style="margin-top:1rem"><a href="/verein/datenschutz">Datenschutzerklärung</a> · <a href="/verein/nutzungsbedingungen">Nutzungsbedingungen</a></p>"""
    return _page(f"Dashboard – {verein_name}", body)


# ── Neuer Termin ─────────────────────────────────────────────────────────────

@verein_bp.route("/verein/termine/neu", methods=["GET", "POST"])
@require_verein_login
def termin_neu(user):
    if user["role"] != "admin":
        return redirect("/verein/dashboard")

    error = ""
    if request.method == "POST":
        datum = request.form.get("datum", "").strip()
        uhrzeit = request.form.get("uhrzeit", "").strip()
        bezeichnung = request.form.get("bezeichnung", "").strip()
        ort = request.form.get("ort", "").strip()

        if not datum or not bezeichnung:
            error = "Datum und Bezeichnung sind Pflichtfelder."
        else:
            termin_id = str(uuid.uuid4())[:8]
            neuer_termin = {
                "id": termin_id,
                "datum": datum,
                "uhrzeit": uhrzeit,
                "bezeichnung": bezeichnung,
                "ort": ort,
                "erstellt_von": user["email"],
                "erstellt_am": datetime.utcnow().isoformat()[:19],
            }
            verein_key = user["verein_key"]

            def updater(data):
                if verein_key not in data:
                    data[verein_key] = []
                    data["_labels"] = data.get("_labels", {})
                    data["_labels"][verein_key] = user["verein_name"]
                data[verein_key].append(neuer_termin)
                data.setdefault("_meta", {}).setdefault(verein_key, {})["selbstverwaltung"] = True
                return data

            KalenderStore.update(updater)
            log_audit("erstellt", termin_id, verein_key, user["id"])
            return redirect("/verein/dashboard")

    form = f"""
{'<p class="err">'+error+'</p>' if error else ''}
<form method="post" autocomplete="off">
  <label>Datum *</label>
  <input name="datum" type="date" required value="{date.today().isoformat()}">
  <label>Uhrzeit</label>
  <input name="uhrzeit" type="time" placeholder="optional">
  <label>Bezeichnung *</label>
  <input name="bezeichnung" type="text" required placeholder="z.B. Jahreshauptversammlung">
  <label>Ort / Veranstaltungsort</label>
  <input name="ort" type="text" placeholder="z.B. Gasthaus zur Post">
  <button class="btn" type="submit">Termin speichern</button>
</form>
{_BACK_DASH}"""
    return _page("Neuer Termin", form)


# ── Termin bearbeiten / löschen ───────────────────────────────────────────────

@verein_bp.route("/verein/termine/<termin_id>", methods=["GET", "POST"])
@require_verein_login
def termin_edit(user, termin_id):
    if user["role"] != "admin":
        return redirect("/verein/dashboard")

    verein_key = user["verein_key"]
    data = _load_data()
    termine_list = data.get(verein_key, [])
    termin = next((t for t in termine_list if t.get("id") == termin_id and not t.get("deleted")), None)

    if not termin:
        return redirect("/verein/dashboard")

    if request.method == "POST":
        aktion = request.form.get("aktion", "")
        if aktion == "loeschen":
            def del_updater(d):
                for t in d.get(verein_key, []):
                    if t.get("id") == termin_id:
                        t["deleted"] = True
                        t["geloescht_von"] = user["email"]
                        t["geloescht_am"] = datetime.utcnow().isoformat()[:19]
                return d
            KalenderStore.update(del_updater)
            log_audit("geloescht", termin_id, verein_key, user["id"])
            return redirect("/verein/dashboard")
        else:
            datum = request.form.get("datum", "").strip()
            uhrzeit = request.form.get("uhrzeit", "").strip()
            bezeichnung = request.form.get("bezeichnung", "").strip()
            ort = request.form.get("ort", "").strip()
            if datum and bezeichnung:
                def edit_updater(d):
                    for t in d.get(verein_key, []):
                        if t.get("id") == termin_id:
                            t["datum"] = datum
                            t["uhrzeit"] = uhrzeit
                            t["bezeichnung"] = bezeichnung
                            t["ort"] = ort
                            t["geaendert_von"] = user["email"]
                            t["geaendert_am"] = datetime.utcnow().isoformat()[:19]
                    return d
                KalenderStore.update(edit_updater)
                log_audit("geaendert", termin_id, verein_key, user["id"])
                return redirect("/verein/dashboard")

    form = f"""
<form method="post">
  <label>Datum</label>
  <input name="datum" type="date" required value="{termin.get('datum','')}">
  <label>Uhrzeit</label>
  <input name="uhrzeit" type="time" value="{termin.get('uhrzeit','')}">
  <label>Bezeichnung</label>
  <input name="bezeichnung" type="text" required value="{termin.get('bezeichnung','')}">
  <label>Ort</label>
  <input name="ort" type="text" value="{termin.get('ort','')}">
  <button class="btn" type="submit" name="aktion" value="speichern">Änderungen speichern</button>
</form>
<hr>
<form method="post" onsubmit="return confirm('Termin wirklich löschen?')">
  <button class="btn btn-danger" type="submit" name="aktion" value="loeschen">🗑 Termin löschen</button>
</form>
{_BACK_DASH}"""
    return _page("Termin bearbeiten", form)


# ── Passwort ändern ───────────────────────────────────────────────────────────

@verein_bp.route("/verein/passwort", methods=["GET", "POST"])
@require_verein_login
def change_password(user):
    from services.auth.routes import _hash_pw, _check_pw
    error = ""
    success = ""
    if request.method == "POST":
        alt = request.form.get("password_alt", "")
        neu = request.form.get("password_neu", "")
        neu2 = request.form.get("password_neu2", "")
        with db_conn() as conn:
            row = conn.execute(
                "SELECT password_hash FROM vk_users WHERE id=?", (user["id"],)
            ).fetchone()
            if not _check_pw(alt, row["password_hash"]):
                error = "Aktuelles Passwort falsch."
            elif len(neu) < 8:
                error = "Neues Passwort muss mindestens 8 Zeichen haben."
            elif neu != neu2:
                error = "Neue Passwörter stimmen nicht überein."
            else:
                conn.execute(
                    "UPDATE vk_users SET password_hash=? WHERE id=?",
                    (_hash_pw(neu), user["id"]),
                )
                success = "Passwort erfolgreich geändert."

    form = f"""
{'<p class="err">'+error+'</p>' if error else ''}
{'<p class="ok">'+success+'</p>' if success else ''}
<form method="post" autocomplete="off">
  <label>Aktuelles Passwort</label>
  <input name="password_alt" type="password" required autocomplete="current-password">
  <label>Neues Passwort</label>
  <input name="password_neu" type="password" required autocomplete="new-password">
  <label>Neues Passwort wiederholen</label>
  <input name="password_neu2" type="password" required autocomplete="new-password">
  <button class="btn" type="submit">Passwort ändern</button>
</form>
{_BACK_DASH}"""
    return _page("Passwort ändern", form)


# ── Mitglieder ────────────────────────────────────────────────────────────────

@verein_bp.route("/verein/mitglieder", methods=["GET", "POST"])
@require_verein_login
def mitglieder(user):
    if user["role"] != "admin":
        return redirect("/verein/dashboard")

    import secrets as _sec
    from shared.vk_mail import send_invite_email

    error = ""
    success = ""
    if request.method == "POST":
        aktion = request.form.get("aktion", "")
        if aktion == "einladen":
            email = request.form.get("email", "").strip().lower()
            with db_conn() as conn:
                count = conn.execute(
                    "SELECT COUNT(*) FROM vk_users WHERE verein_id=?", (user["verein_id"],)
                ).fetchone()[0]
                if count >= 3:
                    error = "Maximal 3 Accounts pro Verein (Admin + 2 Mitglieder)."
                else:
                    ex = conn.execute(
                        "SELECT id FROM vk_users WHERE email=?", (email,)
                    ).fetchone()
                    if ex:
                        error = "Diese E-Mail ist bereits registriert."
                    else:
                        import bcrypt as _bc
                        token = _sec.token_urlsafe(32)
                        expires = (datetime.utcnow() + timedelta(hours=48)).isoformat()
                        tmp_hash = _bc.hashpw(_sec.token_hex(16).encode(), _bc.gensalt()).decode()
                        conn.execute(
                            """INSERT INTO vk_users
                               (email, password_hash, verein_id, role,
                                einladungs_token, einladungs_expires, email_verified, aktiv)
                               VALUES (?,?,?,'member',?,?,1,0)""",
                            (email, tmp_hash, user["verein_id"], token, expires),
                        )
                        send_invite_email(email, token, user["verein_name"])
                        success = f"Einladung an {email} verschickt."
        elif aktion == "entfernen":
            member_id = int(request.form.get("member_id", 0))
            with db_conn() as conn:
                conn.execute(
                    "DELETE FROM vk_users WHERE id=? AND verein_id=? AND role='member'",
                    (member_id, user["verein_id"]),
                )
                success = "Mitglied entfernt."

    with db_conn() as conn:
        members = conn.execute(
            "SELECT id, email, role, aktiv, email_verified FROM vk_users WHERE verein_id=? ORDER BY id",
            (user["verein_id"],),
        ).fetchall()

    rows = ""
    for m in members:
        status = "✅" if m["aktiv"] else "⏳ Einladung ausstehend"
        remove_btn = ""
        if m["role"] == "member":
            remove_btn = f'<form method="post" style="display:inline"><input type="hidden" name="aktion" value="entfernen"><input type="hidden" name="member_id" value="{m["id"]}"><button style="background:none;border:none;color:#ff453a;cursor:pointer;font-size:.9rem" type="submit">Entfernen</button></form>'
        rows += f'<div class="card"><div style="display:flex;justify-content:space-between"><div><div>{m["email"]}</div><div style="color:#aeaeb2;font-size:.82rem">{m["role"]} · {status}</div></div><div>{remove_btn}</div></div></div>'

    invite_form = ""
    if len(members) < 3:
        invite_form = f"""
{'<p class="err">'+error+'</p>' if error else ''}
{'<p class="ok">'+success+'</p>' if success else ''}
<form method="post">
  <input type="hidden" name="aktion" value="einladen">
  <label>E-Mail-Adresse des neuen Mitglieds</label>
  <input name="email" type="email" required placeholder="mitglied@beispiel.de">
  <button class="btn" type="submit">Einladung verschicken</button>
</form>
<div class="spam-hint">📬 Bitte Eingeladene auf den Spam-Ordner hinweisen.</div>"""
    else:
        invite_form = '<p class="hint">Maximale Anzahl (3) erreicht.</p>'
        if success:
            invite_form = f'<p class="ok">{success}</p>' + invite_form

    body = f"""
<h2 style="font-size:1rem;margin:0 0 .5rem">Mitglieder ({len(members)}/3)</h2>
{rows}
<hr>
{invite_form}
{_BACK_DASH}"""
    return _page("Mitglieder", body)


# ── Einladung annehmen ────────────────────────────────────────────────────────

@verein_bp.route("/verein/einladung", methods=["GET", "POST"])
def einladung():
    from services.auth.routes import _hash_pw
    token = request.args.get("token", "") or request.form.get("token", "")
    error = ""
    with db_conn() as conn:
        row = conn.execute(
            "SELECT id, einladungs_expires, aktiv FROM vk_users WHERE einladungs_token=?",
            (token,),
        ).fetchone()
        if not row:
            return _page("Ungültig", '<p class="err">Ungültiger Einladungslink.</p>'), 400
        if datetime.fromisoformat(row["einladungs_expires"]) < datetime.utcnow():
            return _page("Abgelaufen", '<p class="err">Der Einladungslink ist abgelaufen. Bitte den Vereinsadmin um eine neue Einladung bitten.</p>'), 400
        if row["aktiv"]:
            return redirect("/verein/login")

        if request.method == "POST":
            pw = request.form.get("password", "")
            pw2 = request.form.get("password2", "")
            if len(pw) < 8:
                error = "Passwort muss mindestens 8 Zeichen haben."
            elif pw != pw2:
                error = "Passwörter stimmen nicht überein."
            else:
                conn.execute(
                    """UPDATE vk_users SET password_hash=?, aktiv=1,
                       einladungs_token=NULL, einladungs_expires=NULL WHERE id=?""",
                    (_hash_pw(pw), row["id"]),
                )
                return redirect("/verein/login")

    form = f"""
<p>Lege dein Passwort fest, um die Einladung anzunehmen.</p>
{'<p class="err">'+error+'</p>' if error else ''}
<form method="post">
  <input type="hidden" name="token" value="{token}">
  <label>Passwort <span class="hint">(mind. 8 Zeichen)</span></label>
  <input name="password" type="password" required autocomplete="new-password">
  <label>Passwort wiederholen</label>
  <input name="password2" type="password" required autocomplete="new-password">
  <button class="btn" type="submit">Einladung annehmen</button>
</form>"""
    return _page("Einladung annehmen", form)


# ── Upload-Template ──────────────────────────────────────────────────────────

@verein_bp.route("/verein/upload-template")
@require_verein_login
def upload_template(user):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Termine"

    header_labels = ["Datum *", "Uhrzeit", "Bezeichnung *", "Ort", "Ortschaft"]
    hfill = PatternFill("solid", fgColor="6D28D9")
    hfont = Font(bold=True, color="FFFFFF")
    for col, label in enumerate(header_labels, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center")

    for row_data in [
        ("15.06.2026", "19:00", "Jahreshauptversammlung", "GH Zur Post, Hölskofen", "Hölskofen"),
        ("22.06.2026", "",      "Sommerfest",             "Festplatz Postau",       "Postau"),
    ]:
        ws.append(row_data)

    for col, w in zip("ABCDE", [14, 10, 35, 30, 18]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = "attachment; filename=terminplan-vorlage.xlsx"
    return resp


# ── Upload-Seite (GET) ────────────────────────────────────────────────────────

@verein_bp.route("/verein/upload", methods=["GET"])
@require_verein_login
def upload_page(user):
    if user["role"] != "admin":
        return redirect("/verein/dashboard")

    remaining = _quota_remaining(user["verein_id"])
    used = _UPLOAD_LIMIT - remaining

    if remaining == 0:
        body = (
            f'<p class="err">Tageslimit erreicht ({_UPLOAD_LIMIT}/{_UPLOAD_LIMIT} Uploads heute). '
            f'Morgen wieder verfügbar.</p>{_BACK_DASH}'
        )
        return _page("Terminplan hochladen", body)

    quota_bar = (
        f'<p class="hint" style="margin-bottom:1rem">'
        f'Uploads heute: {used}/{_UPLOAD_LIMIT}</p>'
    )
    body = f"""{quota_bar}
<div class="card">
  <h2 style="font-size:.95rem;margin-top:0">📄 PDF oder Foto</h2>
  <p class="hint">Claude KI extrahiert die Termine automatisch aus dem Dokument.</p>
  <form method="post" action="/verein/upload" enctype="multipart/form-data">
    <input type="hidden" name="typ" value="vision">
    <label>Datei (PDF, JPG, PNG, HEIC)</label>
    <input type="file" name="file" required accept=".pdf,.jpg,.jpeg,.png,.heic,.heif">
    <button class="btn" type="submit">Hochladen &amp; analysieren</button>
  </form>
</div>
<div class="card">
  <h2 style="font-size:.95rem;margin-top:0">📊 Excel-Tabelle</h2>
  <p class="hint">Trage Termine in die Vorlage ein und lade sie hoch – ohne KI, keine Extraktion.</p>
  <a class="btn btn-sec" href="/verein/upload-template"
     style="margin-bottom:.75rem">⬇ Vorlage herunterladen (.xlsx)</a>
  <form method="post" action="/verein/upload" enctype="multipart/form-data">
    <input type="hidden" name="typ" value="excel">
    <label>Ausgefüllte Excel-Datei (.xlsx)</label>
    <input type="file" name="file" required accept=".xlsx">
    <button class="btn" type="submit">Termine importieren</button>
  </form>
</div>
{_BACK_DASH}"""
    return _page(f"Terminplan hochladen – {user['verein_name']}", body)


# ── Upload verarbeiten (POST) ─────────────────────────────────────────────────

@verein_bp.route("/verein/upload", methods=["POST"])
@require_verein_login
def upload_process(user):
    if user["role"] != "admin":
        return redirect("/verein/dashboard")

    verein_name = user["verein_name"]
    verein_key  = user["verein_key"]
    verein_id   = user["verein_id"]
    heute       = date.today().isoformat()

    if _quota_remaining(verein_id) <= 0:
        body = (
            f'<p class="err">Tageslimit erreicht ({_UPLOAD_LIMIT}/{_UPLOAD_LIMIT} Uploads heute).</p>'
            + _BACK_DASH
        )
        return _page("Limit erreicht", body), 429

    if "file" not in request.files:
        return redirect("/verein/upload")

    f      = request.files["file"]
    fname  = (f.filename or "").lower()
    suffix = Path(fname).suffix if fname else ""
    typ    = request.form.get("typ", "vision")

    # Quota jetzt erhöhen (vor dem Claude-Call, um Missbrauch zu verhindern)
    increment_upload_quota(verein_id, heute)

    auto_plz = ""
    alle: list = []

    if typ == "excel":
        if suffix != ".xlsx":
            body = '<p class="err">Nur .xlsx-Dateien erlaubt.</p>' + _BACK_DASH
            return _page("Fehler", body), 400
        try:
            alle = parse_excel_bytes(f.read())
        except Exception as ex:
            body = f'<p class="err">Fehler beim Lesen der Excel-Datei: {html.escape(str(ex))}</p>' + _BACK_DASH
            return _page("Fehler", body), 400
    else:
        allowed = {".pdf", ".jpg", ".jpeg", ".png", ".heic", ".heif"}
        if suffix not in allowed:
            body = '<p class="err">Nur PDF oder Bilder (JPG, PNG, HEIC) erlaubt.</p>' + _BACK_DASH
            return _page("Fehler", body), 400
        if suffix in {".heic", ".heif"} and not _HEIC_SUPPORTED:
            body = '<p class="err">HEIC-Format auf diesem Server nicht verfügbar.</p>' + _BACK_DASH
            return _page("Fehler", body), 400
        try:
            result   = import_pdf_bytes(f.read(), suffix)
            alle     = result["alle"]
            auto_plz = result.get("auto_plz", "")
        except Exception as ex:
            body = f'<p class="err">Fehler bei der KI-Analyse: {html.escape(str(ex))}</p>' + _BACK_DASH
            return _page("Fehler", body), 500

    if not alle:
        body = '<p class="err">Keine Termine gefunden.</p>' + _BACK_DASH
        return _page("Keine Termine", body)

    # Alle Termine diesem Verein zuordnen
    for t in alle:
        t["verein"] = verein_name
        t["quelle"] = verein_name

    try:
        data = json.loads(VEREINSTERMINE_FILE.read_text()) if VEREINSTERMINE_FILE.exists() else {}
    except Exception:
        data = {}

    ort_cfg     = data.get("_ortschaften", {"whitelist": [], "blacklist": []})
    known_white = set(ort_cfg.get("whitelist", []))
    known_black = set(ort_cfg.get("blacklist", []))
    neue_orts   = sorted({
        t.get("ortschaft", "").strip()
        for t in alle
        if t.get("ortschaft", "").strip()
           and t["ortschaft"].strip() not in known_white
           and t["ortschaft"].strip() not in known_black
    })

    if neue_orts:
        import_id = str(uuid.uuid4())
        Path(f"/tmp/vk_pending_{import_id}.json").write_text(
            json.dumps({
                "import_id":   import_id,
                "alle":        alle,
                "auto_plz":    auto_plz,
                "form_plz":    "",
                "verein_id":   verein_id,
                "verein_key":  verein_key,
                "verein_name": verein_name,
            }, ensure_ascii=False)
        )

        ort_items = ""
        for o in neue_orts:
            o_esc = html.escape(o)
            ort_items += f"""
<div class="card" style="padding:.75rem 1rem">
  <div style="font-weight:500;margin-bottom:.4rem">{o_esc}</div>
  <input type="hidden" name="alle_orts" value="{o_esc}">
  <label style="display:flex;align-items:center;gap:.5rem;margin:0;font-size:.9rem">
    <input type="checkbox" name="confirm" value="{o_esc}" checked>
    In Ortsliste aufnehmen
  </label>
</div>"""

        preview = sorted(alle, key=lambda x: x.get("datum", ""))
        prev_rows = "".join(
            f'<div style="font-size:.82rem;color:#aeaeb2;padding:.3rem 0;'
            f'border-bottom:1px solid #3a3a3c">'
            f'{html.escape(t.get("datum",""))} – {html.escape(t.get("bezeichnung",""))}'
            + (f' · {html.escape(t.get("ortschaft",""))}' if t.get("ortschaft") else "")
            + '</div>'
            for t in preview[:10]
        )
        more = f'<p class="hint">… und {len(alle) - 10} weitere</p>' if len(alle) > 10 else ""

        body = f"""
<p class="hint">{len(alle)} Termine gefunden · {len(neue_orts)} neue Ortschaft(en) prüfen</p>
<h2 style="font-size:.95rem;margin:1rem 0 .5rem">Neue Ortschaften</h2>
<form method="post" action="/verein/confirm-upload">
  <input type="hidden" name="import_id" value="{import_id}">
  {ort_items}
  <h2 style="font-size:.95rem;margin:1.25rem 0 .5rem">Vorschau (erste 10)</h2>
  <div class="card">{prev_rows}{more}</div>
  <button class="btn" type="submit" style="margin-top:1rem">Termine importieren</button>
</form>
{_BACK_DASH}"""
        return _page("Ortschaften prüfen", body)

    # Keine neuen Ortschaften – direkt speichern
    _, total = _do_save_import(alle, auto_plz, "", data)
    def _sv_up(d): d.setdefault("_meta", {}).setdefault(verein_key, {})["selbstverwaltung"] = True; return d
    KalenderStore.update(_sv_up)
    log_audit("upload", f"bulk_{total}", verein_key, user["id"])
    return redirect(f"/verein/dashboard?upload_ok={total}")


# ── Upload bestätigen (POST) ──────────────────────────────────────────────────

@verein_bp.route("/verein/confirm-upload", methods=["POST"])
@require_verein_login
def confirm_upload(user):
    if user["role"] != "admin":
        return redirect("/verein/dashboard")

    import_id    = request.form.get("import_id", "")
    pending_path = Path(f"/tmp/vk_pending_{import_id}.json")

    if not pending_path.exists():
        body = (
            '<p class="err">Import nicht gefunden oder abgelaufen. '
            'Bitte Datei erneut hochladen.</p>' + _BACK_DASH
        )
        return _page("Fehler", body), 404

    pending = json.loads(pending_path.read_text())

    if pending.get("verein_id") != user["verein_id"]:
        body = '<p class="err">Nicht autorisiert.</p>' + _BACK_DASH
        return _page("Fehler", body), 403

    alle_orts    = request.form.getlist("alle_orts")
    confirm_list = request.form.getlist("confirm")
    reject_list  = [o for o in alle_orts if o not in confirm_list]

    try:
        data = json.loads(VEREINSTERMINE_FILE.read_text()) if VEREINSTERMINE_FILE.exists() else {}
    except Exception:
        data = {}

    ort_cfg   = data.get("_ortschaften", {"whitelist": [], "blacklist": []})
    whitelist = set(ort_cfg.get("whitelist", []))
    blacklist = set(ort_cfg.get("blacklist", []))
    for o in confirm_list:
        whitelist.add(o); blacklist.discard(o)
    for o in reject_list:
        blacklist.add(o); whitelist.discard(o)
    data["_ortschaften"] = {"whitelist": sorted(whitelist), "blacklist": sorted(blacklist)}

    vk = user["verein_key"]
    _, total = _do_save_import(pending["alle"], pending.get("auto_plz", ""), "", data)
    pending_path.unlink(missing_ok=True)
    def _sv_cf(d): d.setdefault("_meta", {}).setdefault(vk, {})["selbstverwaltung"] = True; return d
    KalenderStore.update(_sv_cf)
    log_audit("upload_confirmed", f"bulk_{total}", vk, user["id"])
    return redirect(f"/verein/dashboard?upload_ok={total}")


# ── Datenschutz / Nutzungsbedingungen ────────────────────────────────────────

@verein_bp.route("/verein/datenschutz")
def datenschutz():
    body = """
<p style="color:#aeaeb2;font-size:.85rem">Stand: Mai 2026</p>
<div class="card">
<h2 style="font-size:1rem;margin-top:0">1. Verantwortlicher</h2>
<p>Josef Fischer, Hölskofen 13, 84092 Bayerbach b. Ergoldsbach · <a href="mailto:Vereinskalender@icloud.com">Vereinskalender@icloud.com</a></p>
</div>
<div class="card">
<h2 style="font-size:1rem;margin-top:0">2. Erhobene Daten</h2>
<p>Bei der Registrierung erfassen wir: E-Mail-Adresse, Vereinsname sowie das Passwort (verschlüsselt gespeichert, niemals im Klartext). Vereinstermine werden öffentlich angezeigt.</p>
</div>
<div class="card">
<h2 style="font-size:1rem;margin-top:0">3. Zweck der Verarbeitung</h2>
<p>Die Daten dienen ausschließlich dem Betrieb des Vereinskalenders: Identifizierung des Vereins, Authentifizierung des Accounts und Benachrichtigungen (Bestätigungs-E-Mails).</p>
</div>
<div class="card">
<h2 style="font-size:1rem;margin-top:0">4. Speicherdauer</h2>
<p>Accounts werden auf Anfrage gelöscht. Schreib dazu an <a href="mailto:Vereinskalender@icloud.com">Vereinskalender@icloud.com</a>. Eingetragene Termine werden nach Ende des jeweiligen Kalenderjahres bereinigt.</p>
</div>
<div class="card">
<h2 style="font-size:1rem;margin-top:0">5. E-Mail-Dienst</h2>
<p>E-Mails werden über Brevo (Sendinblue SAS, Frankreich) versendet. Dabei wird die Ziel-E-Mail-Adresse an Brevo übermittelt.</p>
</div>
<div class="card">
<h2 style="font-size:1rem;margin-top:0">5a. Telegram-Terminerinnerungen (freiwillig)</h2>
<p>Wer den Telegram-Bot für Terminerinnerungen nutzt, speichert damit freiwillig seine Telegram-Chat-ID sowie die ausgewählten Vereins-Abonnements auf unserem Server. Diese Daten werden ausschließlich zum Versand der gewünschten Erinnerungen verwendet. Abmelden ist jederzeit mit dem Befehl /stop im Bot möglich – dabei werden alle gespeicherten Daten gelöscht. Eine Löschung ist auch per E-Mail möglich.</p>
</div>
<div class="card">
<h2 style="font-size:1rem;margin-top:0">6. Rechte</h2>
<p>Auskunft, Berichtigung, Löschung deiner Daten: Schreib an <a href="mailto:Vereinskalender@icloud.com">Vereinskalender@icloud.com</a>. Beschwerderecht bei der zuständigen Datenschutz-Aufsichtsbehörde.</p>
</div>
{_BACK_DASH}"""
    return _page("Datenschutzerklärung", body)


@verein_bp.route("/verein/nutzungsbedingungen")
def nutzungsbedingungen():
    body = """
<p style="color:#aeaeb2;font-size:.85rem">Stand: Mai 2026</p>
<div class="card">
<h2 style="font-size:1rem;margin-top:0">1. Nutzung</h2>
<p>Der Vereinskalender vereinskalender.online dient der nicht-kommerziellen Veröffentlichung von Vereinsterminen. Die Nutzung ist kostenlos.</p>
</div>
<div class="card">
<h2 style="font-size:1rem;margin-top:0">2. Registrierung</h2>
<p>Nur Vereine im Einzugsgebiet dürfen sich registrieren. Gewerbliche oder kommerzielle Anbieter sind ausgeschlossen. Jeder Verein trägt Verantwortung für die Richtigkeit seiner Daten.</p>
</div>
<div class="card">
<h2 style="font-size:1rem;margin-top:0">3. Haftungsausschluss</h2>
<p>Der Betreiber übernimmt keine Gewähr für die Richtigkeit der eingetragenen Termine. Urheberrechtlich geschützte Inhalte dürfen nicht ohne Genehmigung eingestellt werden.</p>
</div>
<div class="card">
<h2 style="font-size:1rem;margin-top:0">4. Kündigung</h2>
<p>Der Betreiber kann Accounts bei Verstoß gegen diese Bedingungen ohne Vorankündigung sperren oder löschen.</p>
</div>
{_BACK_DASH}"""
    return _page("Nutzungsbedingungen", body)


# ── Vereinsprofil ─────────────────────────────────────────────────────────────

@verein_bp.route("/verein/profil", methods=["GET", "POST"])
@require_verein_login
def verein_profil(user):
    if user["role"] != "admin":
        return redirect("/verein/dashboard")

    error = ok = ""

    with db_conn() as conn:
        va = conn.execute(
            "SELECT verein_name, rubrik, heimatort, plz, gemeinde, landkreis FROM vereine_accounts WHERE id=?",
            (user["verein_id"],),
        ).fetchone()
        usr = conn.execute(
            "SELECT telefon FROM vk_users WHERE id=?", (user["id"],)
        ).fetchone()

    if not va:
        return redirect("/verein/dashboard")

    verein_name = va["verein_name"]
    rubrik      = va["rubrik"] or "Verein"
    heimatort   = va["heimatort"] or ""
    plz         = va["plz"] or ""
    gemeinde    = va["gemeinde"] or ""
    landkreis   = va["landkreis"] or ""
    telefon     = usr["telefon"] if usr else ""

    if request.method == "POST":
        new_name      = request.form.get("verein_name", "").strip()
        new_rubrik    = request.form.get("rubrik", "").strip()
        new_heimatort = request.form.get("heimatort", "").strip()
        new_plz       = request.form.get("plz", "").strip()
        new_telefon   = request.form.get("telefon", "").strip()

        if not new_name or len(new_name) < 3:
            error = "Vereinsname muss mindestens 3 Zeichen haben."
        elif new_rubrik not in RUBRIKEN:
            error = "Bitte eine gültige Rubrik wählen."
        elif not new_heimatort or len(new_heimatort) < 2:
            error = "Bitte einen Heimatort angeben."
        elif new_plz and not re.match(r"^\d{5}$", new_plz):
            error = "PLZ muss 5 Ziffern haben."
        else:
            new_gemeinde = new_landkreis = ""
            if new_plz and new_plz != plz:
                geo = lookup_plz(new_plz)
                new_gemeinde  = geo.get("gemeinde", "")
                new_landkreis = geo.get("landkreis", "")
            elif not new_plz:
                new_gemeinde = new_landkreis = ""
            else:
                new_gemeinde  = gemeinde
                new_landkreis = landkreis

            with db_conn() as conn:
                conn.execute(
                    """UPDATE vereine_accounts
                       SET verein_name=?, rubrik=?, heimatort=?, plz=?, gemeinde=?, landkreis=?
                       WHERE id=?""",
                    (new_name, new_rubrik, new_heimatort,
                     new_plz or None, new_gemeinde or None, new_landkreis or None,
                     user["verein_id"]),
                )
                conn.execute(
                    "UPDATE vk_users SET telefon=? WHERE id=?",
                    (new_telefon or None, user["id"]),
                )
            verein_name = new_name
            rubrik      = new_rubrik
            heimatort   = new_heimatort
            plz         = new_plz
            gemeinde    = new_gemeinde
            landkreis   = new_landkreis
            telefon     = new_telefon
            ok = "✅ Profil gespeichert."

    rubrik_opts = "".join(
        f'<option value="{r}"{" selected" if rubrik == r else ""}>{r}</option>'
        for r in RUBRIKEN
    )
    geo_hint = f'<p class="hint">{gemeinde}, {landkreis}</p>' if gemeinde else ""

    body = f"""
{'<p class="err">'+error+'</p>' if error else ''}
{'<p class="ok">'+ok+'</p>' if ok else ''}
<form method="post">
  <label>Vereinsname</label>
  <input name="verein_name" type="text" required value="{html.escape(verein_name)}">
  <label>Rubrik</label>
  <select name="rubrik" required>
    {rubrik_opts}
  </select>
  <label>Heimatort</label>
  <input name="heimatort" type="text" required placeholder="z.B. Musterdorf" value="{html.escape(heimatort)}">
  <label>PLZ <span class="hint">(optional)</span></label>
  <input name="plz" type="text" inputmode="numeric" maxlength="5" placeholder="z.B. 83308" value="{html.escape(plz)}">
  {geo_hint}
  <label>Telefon Ansprechpartner <span class="hint">(optional)</span></label>
  <input name="telefon" type="tel" autocomplete="tel" placeholder="z.B. 0172 1234567" value="{html.escape(telefon or '')}">
  <button class="btn" type="submit">Speichern</button>
</form>
{_BACK_DASH}"""
    return _page("Vereinsprofil", body)
