#!/opt/rename-webhook/bin/python3
"""
heimat_import.py
Fetcht Events von heimat-info.de für alle konfigurierten Gemeinden
und sendet Telegram-Vorschau mit ✅/❌ Import-Buttons.

Aufruf:
  python3 heimat_import.py              → Import-Vorschau per Telegram
  python3 heimat_import.py --add <url>  → Neue Gemeinde via Playwright entdecken
"""
import html as htmlmod
import json
import re
import sys
import urllib.parse
import urllib.request
import uuid
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, "/opt/rename-webhook")
from shared.secrets import load_secrets
from shared.telegram import send_telegram, send_telegram_inline

GEMEINDEN_FILE      = Path("/opt/rename-webhook/heimat_gemeinden.json")
VEREINSTERMINE_FILE = Path("/opt/rename-webhook/vereinstermine.json")
PENDING_DIR         = Path("/opt/rename-webhook/imports")
LOG_FILE            = "/var/log/pka-heimat.log"
API_BASE            = "https://www.heimat-info.de/embeddings/events/v1/"
API_EXPORT          = "https://heimatinfo-api-platform.azurewebsites.net"
API_EXPORT_HEADERS  = {
    "Origin":     "https://www.heimat-info.de",
    "Referer":    "https://www.heimat-info.de/",
    "User-Agent": "Mozilla/5.0",
}

_org_cache: dict[str, str] = {}

MONATE     = {"Januar":1,"Februar":2,"März":3,"April":4,"Mai":5,"Juni":6,
               "Juli":7,"August":8,"September":9,"Oktober":10,"November":11,"Dezember":12}
KATEGORIEN = {"Vereine","Kirchen","Feuerwehren","Gastro / Gewerbe",
               "Veranstaltungen","Sport","Kultur","Sonstiges","Gemeinde","Freizeit"}
WOCHENTAGE = {"So.","Sa.","Mo.","Di.","Mi.","Do.","Fr."}
SKIP_TEXT  = {"zum Kalender hinzufügen","mehr anzeigen"}


def _slugify(name: str) -> str:
    name = name.lower()
    for a, b in [("ä","ae"),("ö","oe"),("ü","ue"),("ß","ss")]:
        name = name.replace(a, b)
    name = re.sub(r"[^a-z0-9]+", "_", name)
    return name.strip("_")[:50]


def _log(msg: str) -> None:
    ts   = datetime.now().isoformat(timespec="seconds")
    line = f"{ts} {msg}"
    print(line)
    with open(LOG_FILE, "a") as f:
        f.write(line + "\n")


def _parse_events(html_content: str, heute: str) -> list[dict]:
    events     = []
    events_raw = re.findall(
        r'<div class="event mb-3"[^>]*>(.*?)(?=<div class="event mb-3"|$)',
        html_content, re.S)

    for e in events_raw:
        m_tag   = re.search(r'class="date-number"[^>]*>(\d+)<', e)
        m_monat = re.search(r'class="date-month"[^>]*>(\w+)<',  e)
        m_jahr  = re.search(r'class="date-year"[^>]*>(\d{4})<', e)
        if not (m_tag and m_monat and m_jahr):
            continue
        monat_nr = MONATE.get(m_monat.group(1), 0)
        if not monat_nr:
            continue
        datum = f'{m_jahr.group(1)}-{monat_nr:02d}-{int(m_tag.group(1)):02d}'
        if datum < heute:
            continue

        m_uhr   = re.search(r'(\d{2}:\d{2}) Uhr', e)
        uhrzeit = m_uhr.group(1) if m_uhr else ""

        texts = []
        for t in re.findall(r'>([^<>\n]{3,200})<', e):
            t = htmlmod.unescape(t.strip())
            if (not t or t in SKIP_TEXT or t in WOCHENTAGE
                    or t == m_monat.group(1)
                    or re.match(r'^\d{1,4}$', t)
                    or "Uhr" in t):
                continue
            texts.append(t)

        filtered    = [t for t in texts if t not in KATEGORIEN]
        verein      = filtered[0] if filtered else ""
        bezeichnung = filtered[1] if len(filtered) > 1 else ""
        ort         = filtered[2] if len(filtered) > 2 else ""

        if bezeichnung:
            events.append({"datum": datum, "uhrzeit": uhrzeit,
                           "bezeichnung": bezeichnung, "ort": ort,
                           "_verein_name": verein})
    return events


def _fetch(c_id: str) -> str | None:
    try:
        with urllib.request.urlopen(f"{API_BASE}?c={c_id}", timeout=15) as r:
            return r.read().decode("utf-8")
    except Exception as e:
        _log(f"  ❌ Fetch-Fehler c={c_id}: {e}")
        return None


def _fetch_org_name(org_id: str) -> str:
    """Holt Vereinsname via Organization-API (gecacht)."""
    if org_id in _org_cache:
        return _org_cache[org_id]
    try:
        req = urllib.request.Request(
            f"{API_EXPORT}/organizations/{org_id}", headers=API_EXPORT_HEADERS)
        with urllib.request.urlopen(req, timeout=8) as r:
            name = json.loads(r.read()).get("name", "").strip()
    except Exception:
        name = ""
    _org_cache[org_id] = name
    return name


def _fetch_all_events(c_id: str) -> list[dict]:
    """Holt alle Events via Export-API (pageSize=50 max, paginiert via pageIndex)."""
    all_events, page = [], 0
    while True:
        url = f"{API_EXPORT}/export/events?pageIndex={page}&pageSize=50&c={c_id}"
        try:
            req = urllib.request.Request(url, headers=API_EXPORT_HEADERS)
            with urllib.request.urlopen(req, timeout=15) as r:
                batch = json.loads(r.read())
        except Exception as e:
            _log(f"  ❌ Export-API Fehler page={page} c={c_id[:8]}: {e}")
            break
        if not batch:
            break
        all_events.extend(batch)
        if len(batch) < 50:
            break
        page += 1
    return all_events


def _parse_api_events(api_events: list[dict], heute: str) -> list[dict]:
    """Konvertiert Export-API JSON-Events in internes Format."""
    try:
        from zoneinfo import ZoneInfo
        berlin = ZoneInfo("Europe/Berlin")
    except ImportError:
        berlin = None

    events = []
    for e in api_events:
        if e.get("status") != "Published":
            continue
        start = e.get("startDate") or ""
        if not start:
            continue
        try:
            dt_utc = datetime.strptime(start[:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=timezone.utc)
            if berlin:
                dt_loc = dt_utc.astimezone(berlin)
            else:
                from datetime import timedelta
                dt_loc = dt_utc + timedelta(hours=2)
            datum   = dt_loc.strftime("%Y-%m-%d")
            uhrzeit = "" if start.endswith("T00:00:00Z") else dt_loc.strftime("%H:%M")
        except Exception:
            continue
        if datum < heute:
            continue

        bezeichnung = (e.get("title") or "").strip()
        ort         = (e.get("location") or "").strip()
        org_id      = e.get("organizationId") or ""
        verein_name = _fetch_org_name(org_id) if org_id else ""

        if bezeichnung:
            events.append({
                "datum":        datum,
                "uhrzeit":      uhrzeit,
                "bezeichnung":  bezeichnung,
                "ort":          ort,
                "_verein_name": verein_name,
            })
    return events


def discover_c_id(url: str) -> str | None:
    """Nutzt Playwright einmalig um die c= Gemeinde-ID zu ermitteln.
    Methode 1: Base64-kodierter iframe in .borlabs-hide (neueres Borlabs)
    Methode 2: Netzwerk-Intercept nach Button-Click (älteres Borlabs)
    """
    import base64
    from playwright.sync_api import sync_playwright

    found = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page    = browser.new_page(viewport={"width": 1280, "height": 900})

        def on_request(req):
            if "heimat-info.de/embeddings" in req.url:
                m = re.search(r'c=([a-f0-9\-]{36})', req.url)
                if m:
                    found.append(m.group(1))
        page.on("request", on_request)

        try:
            page.goto(url, wait_until="networkidle", timeout=30000)
            page.wait_for_timeout(3000)

            # Methode 1: Base64-Block aus .borlabs-hide dekodieren
            b64s = page.evaluate(
                '() => [...document.querySelectorAll(".borlabs-hide")].map(d => d.innerText.trim())')
            for b64 in b64s:
                if not b64:
                    continue
                try:
                    decoded = base64.b64decode(b64 + "==").decode("utf-8", errors="ignore")
                    m = re.search(r'c=([a-f0-9\-]{36})', decoded)
                    if m:
                        found.append(m.group(1))
                except Exception:
                    pass

            if not found:
                # Methode 2: Consent-Button klicken → Netzwerk-Intercept
                page.evaluate("""() => {
                    const sels = ['.brlbs-cmpnt-cb-btn','._brlbs-btn-accept-all',
                                  '[class*=\"accept\"]','[class*=\"consent\"]'];
                    for (const s of sels) {
                        const b = document.querySelector(s);
                        if (b) { b.click(); break; }
                    }
                }""")
                page.wait_for_timeout(10000)

        except Exception as e:
            _log(f"  Playwright-Fehler: {e}")
        finally:
            browser.close()
    return found[0] if found else None


def _existing_events(exclude_keys: set | None = None) -> set[tuple[str, str, str]]:
    """Gibt alle (datum, uhrzeit, bezeichnung)-Tripel aus vereinstermine.json zurück.
    exclude_keys: Keys die übersprungen werden (z.B. alte Gemeinde-Keys bei Migration)."""
    if not VEREINSTERMINE_FILE.exists():
        return set()
    data = json.loads(VEREINSTERMINE_FILE.read_text())
    existing = set()
    skip = exclude_keys or set()
    for key, items in data.items():
        if key in skip or not isinstance(items, list):
            continue
        for item in items:
            if isinstance(item, dict) and "datum" in item:
                existing.add((
                    item["datum"],
                    item.get("uhrzeit", ""),
                    item.get("bezeichnung", "").strip().lower(),
                ))
    return existing


def _is_duplicate(datum: str, uhrzeit: str, bezeichnung: str,
                  existing: set[tuple[str, str, str]]) -> bool:
    """Exakter Match + Substring-Check auf gleichem Datum UND gleicher Uhrzeit.
    Zwei Events gleichen Namens zu verschiedenen Zeiten sind keine Duplikate."""
    bez = bezeichnung.strip().lower()
    if (datum, uhrzeit, bez) in existing:
        return True
    if len(bez) < 6:
        return False
    for ex_datum, ex_uhr, ex_bez in existing:
        if ex_datum != datum or ex_uhr != uhrzeit:
            continue
        if bez in ex_bez or ex_bez in bez:
            return True
    return False


def do_import(uid: str, verein_keys: list | None = None) -> str:
    """Schreibt bestätigte Events in vereinstermine.json.
    verein_keys=None: alle Events importieren.
    verein_keys=[...]: nur diese Vereine importieren; Pending-Datei bleibt mit Rest.
    Löscht zuerst alte Gemeinde-Keys (Migration auf per-Veranstalter-Keys)."""
    pending_file = PENDING_DIR / f"heimat_pending_{uid}.json"
    if not pending_file.exists():
        # Fallback für ältere Pending-Dateien in /tmp
        old = Path(f"/tmp/heimat_pending_{uid}.json")
        if old.exists():
            pending_file = old
        else:
            return "⚠️ Pending-Datei nicht gefunden (Server-Neustart?)"

    pending  = json.loads(pending_file.read_text())
    events   = pending["events"]
    filter_keys = set(verein_keys) if verein_keys is not None else None
    data     = json.loads(VEREINSTERMINE_FILE.read_text()) if VEREINSTERMINE_FILE.exists() else {}
    if "_labels" not in data:
        data["_labels"] = {}
    if "_meta" not in data:
        data["_meta"] = {}
    gemeinde_map: dict = data.setdefault("_ortschaften", {}).setdefault("gemeinde_map", {})

    # Alte Gemeinde-Keys entfernen (werden durch per-Veranstalter-Keys ersetzt)
    old_keys: set = set()
    if GEMEINDEN_FILE.exists():
        gemeinden = json.loads(GEMEINDEN_FILE.read_text())
        old_keys  = {g["verein_key"] for g in gemeinden}
        geloescht = [k for k in old_keys if k in data]
        for k in geloescht:
            del data[k]
            data["_labels"].pop(k, None)
            data["_meta"].pop(k, None)
        if geloescht:
            _log(f"🗑 Alte Gemeinde-Keys entfernt: {', '.join(geloescht)}")

    existing = _existing_events(exclude_keys=old_keys)
    neu = duplikat = 0

    for e in events:
        if filter_keys is not None and e["_verein_key"] not in filter_keys:
            continue
        if not e.get("_neu", True):
            duplikat += 1
            continue
        if _is_duplicate(e["datum"], e["uhrzeit"], e["bezeichnung"], existing):
            duplikat += 1
            continue
        key = e["_verein_key"]
        if key not in data:
            data[key] = []
        data["_labels"].setdefault(key, e["_label"])
        verein_gemeinde = gemeinde_map.get(e["_gemeinde"], "")
        # Geo-Felder nur setzen wenn noch kein Eintrag vorhanden (nie überschreiben)
        if key not in data["_meta"]:
            data["_meta"][key] = {
                "heimatort": e["_gemeinde"],
                "gemeinde":  verein_gemeinde,
                "landkreis": e.get("_landkreis", "Landkreis Landshut"),
            }
        ortschaft = e.get("ortschaft", "") or e["_gemeinde"]
        if ortschaft and ortschaft not in gemeinde_map and verein_gemeinde:
            gemeinde_map[ortschaft] = verein_gemeinde
        data[key].append({
            "datum":        e["datum"],
            "uhrzeit":      e["uhrzeit"],
            "bezeichnung":  e["bezeichnung"],
            "veranstalter": e.get("_verein_name", ""),
            "ort":          e["ort"],
            "ortschaft":    ortschaft,
            "quelle":       e.get("quelle", ""),
            "quelle_url":   e.get("quelle_url", ""),
        })
        existing.add((e["datum"], e["uhrzeit"], e["bezeichnung"].strip().lower()))
        neu += 1

    VEREINSTERMINE_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2))
    _log(f"✅ Import: {neu} neu, {duplikat} Duplikate übersprungen")

    # Pending-Datei: bei Teilimport Rest behalten, sonst löschen
    try:
        if filter_keys is not None:
            remaining = [e for e in events if e["_verein_key"] not in filter_keys]
            if remaining:
                pending["events"] = remaining
                pending_file.write_text(json.dumps(pending, ensure_ascii=False))
            else:
                pending_file.unlink(missing_ok=True)
        else:
            pending_file.unlink(missing_ok=True)
    except OSError:
        pass
    return f"✅ {neu} neue Termine importiert, {duplikat} Duplikate übersprungen"


def do_reject(uid: str, verein_keys: list | None = None) -> str:
    """Verwirft Events aus der Pending-Datei.
    verein_keys=None: gesamten Import verwerfen.
    verein_keys=[...]: nur diese Vereine entfernen."""
    pending_file = PENDING_DIR / f"heimat_pending_{uid}.json"
    if not pending_file.exists():
        return "⚠️ Pending-Datei nicht gefunden"
    if verein_keys is None:
        try:
            pending_file.unlink()
        except OSError:
            pass
        return "🗑 Import verworfen"
    pending = json.loads(pending_file.read_text())
    filter_keys = set(verein_keys)
    remaining = [e for e in pending["events"] if e["_verein_key"] not in filter_keys]
    try:
        if remaining:
            pending["events"] = remaining
            pending_file.write_text(json.dumps(pending, ensure_ascii=False))
        else:
            pending_file.unlink()
    except OSError:
        pass
    return f"🗑 {len(verein_keys)} Verein(e) verworfen"


def fetch_and_save_pending(gemeinden_filter: list | None = None) -> dict:
    """Fetcht Events für alle (oder gefilterte) Gemeinden und speichert Pending.
    Gibt Summary-Dict zurück: {uid, neu, duplikate, sv, fehler, gesamt}
    Bei Fehler: {"error": "..."} ohne uid."""
    PENDING_DIR.mkdir(parents=True, exist_ok=True)

    if not GEMEINDEN_FILE.exists():
        return {"error": "heimat_gemeinden.json nicht gefunden"}

    gemeinden = json.loads(GEMEINDEN_FILE.read_text())
    if gemeinden_filter:
        gemeinden = [g for g in gemeinden if g.get("url") in gemeinden_filter
                     or g.get("name") in gemeinden_filter]
    if not gemeinden:
        return {"error": "Keine Gemeinden konfiguriert"}

    heute    = datetime.now().strftime("%Y-%m-%d")
    old_keys = {g["verein_key"] for g in gemeinden}
    existing = _existing_events(exclude_keys=old_keys)

    try:
        _sv_meta = json.loads(VEREINSTERMINE_FILE.read_text()).get("_meta", {}) if VEREINSTERMINE_FILE.exists() else {}
    except Exception:
        _sv_meta = {}

    alle_events: list = []
    fehler: list      = []

    for g in gemeinden:
        _log(f"Fetche {g['name']} (c={g['c_id'][:8]}…)")
        api_events = _fetch_all_events(g["c_id"])
        if not api_events:
            fehler.append(g["name"])
            continue
        events = _parse_api_events(api_events, heute)
        for e in events:
            veranst          = e.get("_verein_name", "")
            e["_verein_key"] = _slugify(veranst) or g["verein_key"]
            e["_label"]      = veranst or g.get("label", g["name"])
            e["_gemeinde"]   = g["name"]
            e["_landkreis"]  = g.get("landkreis", "Landkreis Landshut")
            e["quelle"]      = "heimat-info.de"
            e["quelle_url"]  = g.get("url", "")
            e["_sv"]         = bool(_sv_meta.get(e["_verein_key"], {}).get("selbstverwaltung", False))
            e["_neu"]        = False if e["_sv"] else not _is_duplicate(
                e["datum"], e["uhrzeit"], e["bezeichnung"], existing)
        alle_events.extend(events)
        neu_count = sum(1 for e in events if e["_neu"])
        _log(f"  → {len(events)} Termine ({neu_count} neu)")

    if not alle_events:
        return {"error": "Keine bevorstehenden Termine gefunden", "fehler": fehler}

    alle_events.sort(key=lambda x: (x["datum"], x.get("uhrzeit", "")))
    uid = str(uuid.uuid4())[:8]
    (PENDING_DIR / f"heimat_pending_{uid}.json").write_text(json.dumps({
        "uid":     uid,
        "quelle":  "heimat-info.de",
        "erzeugt": datetime.now().isoformat(timespec="seconds"),
        "events":  alle_events,
    }, ensure_ascii=False))

    neu = sum(1 for e in alle_events if e["_neu"])
    sv  = sum(1 for e in alle_events if e.get("_sv"))
    dup = sum(1 for e in alle_events if not e["_neu"] and not e.get("_sv"))
    _log(f"✅ Pending uid={uid}: {neu} neu, {dup} dup, {sv} sv")
    return {"uid": uid, "neu": neu, "duplikate": dup, "sv": sv,
            "fehler": fehler, "gesamt": len(alle_events)}


def cmd_import(secrets: dict) -> None:
    token   = secrets["TOKEN"]
    chat_id = secrets["CHAT_ID"]

    result = fetch_and_save_pending()
    if "error" in result:
        send_telegram(token, chat_id, f"⚠️ {result['error']}")
        return

    uid        = result["uid"]
    neu_gesamt = result["neu"]
    dup_gesamt = result["duplikate"]
    sv_gesamt  = result["sv"]
    fehler     = result.get("fehler", [])

    pending_file = PENDING_DIR / f"heimat_pending_{uid}.json"
    alle_events  = json.loads(pending_file.read_text())["events"]

    def _vorschau_zeile(e: dict) -> str:
        veranst = e.get("_verein_name", "")
        ort     = e.get("ort", "")
        teile   = [e["bezeichnung"][:30]]
        if veranst: teile.append(veranst[:25])
        if ort:     teile.append(ort[:25])
        return f"• {e['datum']} {e.get('uhrzeit',''):5} – {' · '.join(teile)} [{e['_gemeinde']}]"

    neue     = [e for e in alle_events if e["_neu"]]
    vorschau = "\n".join(_vorschau_zeile(e) for e in neue[:15])
    if len(neue) > 15:
        vorschau += f"\n… +{len(neue)-15} weitere neue"

    sv_labels  = sorted({e.get("_label") or e["_verein_key"] for e in alle_events if e.get("_sv")})
    sv_hinweis = (f"\n\n⛔ Selbstverwaltete Vereine ignoriert ({sv_gesamt} Termine): "
                  + ", ".join(sv_labels)) if sv_gesamt else ""
    zähler     = f"Gesamt: {result['gesamt']} | 🆕 Neu: {neu_gesamt} | ⏭ Duplikate: {dup_gesamt}"
    if sv_gesamt:
        zähler += f" | 🔒 SV: {sv_gesamt}"

    gemeinden     = json.loads(GEMEINDEN_FILE.read_text())
    gemeinden_str = ", ".join(g["name"] for g in gemeinden)
    msg = (f"🏡 heimat-info Import\n"
           f"Gemeinden: {gemeinden_str}\n"
           f"{zähler}\n\n"
           + (vorschau if neue else "Alle Termine bereits vorhanden.")
           + sv_hinweis
           + f"\n\n→ Admin-Bereich: vereinskalender.online/#admin → Importe")

    send_telegram_inline(token, chat_id, msg, [[
        {"text": f"✅ {neu_gesamt} importieren", "callback_data": f"heimat_ok:{uid}"},
        {"text": "❌ Verwerfen",                 "callback_data": f"heimat_no:{uid}"},
    ]])

    if fehler:
        send_telegram(token, chat_id, f"⚠️ Fetch-Fehler bei: {', '.join(fehler)}")


def cmd_add(url: str, secrets: dict) -> None:
    token   = secrets["TOKEN"]
    chat_id = secrets["CHAT_ID"]

    _log(f"Discovery: {url}")
    c_id = discover_c_id(url)

    if not c_id:
        send_telegram(token, chat_id,
                      f"❌ Keine heimat-info ID gefunden auf:\n{url}\n\n"
                      "Prüfe ob die Seite heimat-info nutzt und ein 'Inhalte entsperren'-Button vorhanden ist.")
        return

    # Slug aus URL ableiten
    slug = re.sub(r'https?://(www\.)?', '', url).split('/')[0].replace('gemeinde-', '').replace('.de', '').replace('.', '_')
    slug = re.sub(r'[^a-z0-9_]', '', slug.lower())[:20]
    name = slug.replace('_', ' ').title()

    gemeinden = json.loads(GEMEINDEN_FILE.read_text()) if GEMEINDEN_FILE.exists() else []

    # Duplikat prüfen
    if any(g["c_id"] == c_id for g in gemeinden):
        send_telegram(token, chat_id, f"ℹ️ Diese Gemeinde ist bereits eingetragen (c={c_id[:8]}…)")
        return

    eintrag = {"name": name, "label": name, "verein_key": slug,
                "c_id": c_id, "url": url}
    gemeinden.append(eintrag)
    GEMEINDEN_FILE.write_text(json.dumps(gemeinden, ensure_ascii=False, indent=2))

    # Prüfen ob Gemeinde-Name in gemeinde_map bekannt ist
    try:
        vt = json.loads(VEREINSTERMINE_FILE.read_text()) if VEREINSTERMINE_FILE.exists() else {}
        gmap = vt.get("_ortschaften", {}).get("gemeinde_map", {})
        in_map = name in gmap or any(v == name for v in gmap.values())
    except Exception:
        in_map = True  # im Zweifel keine Warnung

    msg = (f"✅ Gemeinde hinzugefügt:\n"
           f"Name: {name}\nKey: {slug}\nID: {c_id[:8]}…\n\n")
    if not in_map:
        msg += (f"⚠️ '{name}' fehlt noch in der Ortschaft→Gemeinde-Map!\n"
                f"Bitte vor dem ersten /heimat-Import ergänzen:\n"
                f"  Ortschaft '{name}' → offizielle Gemeinde (z.B. 'Gemeinde {name}')\n"
                f"Sonst erscheint kein Gemeinde-Filter-Chip für diese Gemeinde.\n\n")
    msg += "Mit /heimat den ersten Import starten."
    send_telegram(token, chat_id, msg)
    _log(f"✅ Gemeinde hinzugefügt: {name} ({c_id})" + ("" if in_map else " ⚠️ nicht in gemeinde_map"))


def _get_dropbox_token(secrets: dict) -> str:
    data = urllib.parse.urlencode({
        "grant_type":    "refresh_token",
        "refresh_token": secrets["DROPBOX_REFRESH_TOKEN"],
        "client_id":     secrets["DROPBOX_APP_KEY"],
        "client_secret": secrets["DROPBOX_APP_SECRET"],
    }).encode()
    req = urllib.request.Request(
        "https://api.dropbox.com/oauth2/token", data=data, method="POST")
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())["access_token"]


def _upload_dropbox(token: str, file_bytes: bytes, path: str) -> None:
    req = urllib.request.Request(
        "https://content.dropboxapi.com/2/files/upload",
        data=file_bytes, method="POST")
    req.add_header("Authorization",   f"Bearer {token}")
    req.add_header("Content-Type",    "application/octet-stream")
    req.add_header("Dropbox-API-Arg", json.dumps(
        {"path": path, "mode": "overwrite", "mute": True}))
    with urllib.request.urlopen(req, timeout=30):
        pass


def _download_dropbox(token: str, path: str) -> bytes:
    req = urllib.request.Request(
        "https://content.dropboxapi.com/2/files/download", method="POST")
    req.add_header("Authorization",   f"Bearer {token}")
    req.add_header("Dropbox-API-Arg", json.dumps({"path": path}))
    with urllib.request.urlopen(req, timeout=30) as r:
        return r.read()


def _generate_excel(events: list[dict], uid: str) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Vorschau"

    headers = ["importieren", "datum", "uhrzeit", "bezeichnung",
               "veranstalter", "ort", "ortschaft", "gemeinde", "landkreis", "quelle_url"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="6D28D9")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")

    for e in (e for e in events if e.get("_neu", False)):
        ws.append([
            "ja",
            e["datum"],
            e.get("uhrzeit", ""),
            e["bezeichnung"],
            e.get("_verein_name", ""),
            e.get("ort", ""),
            e.get("ortschaft", "") or e.get("_gemeinde", ""),
            e.get("_gemeinde", ""),
            e.get("_landkreis", ""),
            e.get("quelle_url", ""),
        ])

    for i, w in enumerate([12, 12, 8, 40, 30, 25, 20, 15, 20, 40], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws_meta = wb.create_sheet("Meta")
    ws_meta.append(["uid",     uid])
    ws_meta.append(["erzeugt", datetime.now().isoformat(timespec="seconds")])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def cmd_excel(secrets: dict) -> None:
    """Liest bearbeitete Excel von Dropbox, aktualisiert Pending-Datei, sendet neue Vorschau."""
    import io
    from openpyxl import load_workbook

    token   = secrets["TOKEN"]
    chat_id = secrets["CHAT_ID"]

    try:
        db_token = _get_dropbox_token(secrets)
        raw      = _download_dropbox(db_token, DROPBOX_EXCEL_PATH)
    except Exception as exc:
        send_telegram(token, chat_id, f"❌ Dropbox-Download fehlgeschlagen: {exc}")
        return

    try:
        wb     = load_workbook(io.BytesIO(raw))
        ws     = wb["Vorschau"]
        ws_m   = wb["Meta"]
        uid    = ws_m.cell(1, 2).value
    except Exception as exc:
        send_telegram(token, chat_id, f"❌ Excel-Fehler: {exc}")
        return

    if not uid:
        send_telegram(token, chat_id, "❌ UID nicht im Meta-Sheet gefunden.")
        return

    pending_file = PENDING_DIR / f"heimat_pending_{uid}.json"
    if not pending_file.exists():
        send_telegram(token, chat_id,
                      f"⚠️ Pending-Datei nicht gefunden (uid={uid}).\n"
                      "Bitte /heimat erneut ausführen.")
        return

    pending = json.loads(pending_file.read_text())
    events  = pending["events"]

    # Welche Events hat der User auf "ja" gelassen? Key: (datum, uhrzeit, bezeichnung, gemeinde)
    behalten: set[tuple] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        if str(row[0]).strip().lower() == "ja":
            behalten.add((
                str(row[1] or "").strip(),
                str(row[2] or "").strip(),
                str(row[3] or "").strip(),
                str(row[7] or "").strip(),
            ))

    for e in events:
        key    = (e["datum"], e.get("uhrzeit", ""), e["bezeichnung"], e.get("_gemeinde", ""))
        e["_neu"] = key in behalten

    pending_file.write_text(json.dumps(pending, ensure_ascii=False))

    neu_gesamt = sum(1 for e in events if e["_neu"])
    dup_gesamt = len(events) - neu_gesamt
    neue       = [e for e in events if e["_neu"]]

    def _vorschau_zeile(e: dict) -> str:
        veranst = e.get("_verein_name", "")
        ort     = e.get("ort", "")
        teile   = [e["bezeichnung"][:30]]
        if veranst: teile.append(veranst[:25])
        if ort:     teile.append(ort[:25])
        return f"• {e['datum']} {e.get('uhrzeit',''):5} – {' · '.join(teile)} [{e['_gemeinde']}]"

    vorschau = "\n".join(_vorschau_zeile(e) for e in neue[:15])
    if len(neue) > 15:
        vorschau += f"\n… +{len(neue)-15} weitere"

    msg = (f"📊 Excel eingelesen\n"
           f"✅ Zu importieren: {neu_gesamt} | ❌ Ausgeschlossen: {dup_gesamt}\n\n"
           + (vorschau if neue else "Keine Termine zum Importieren ausgewählt."))

    send_telegram_inline(token, chat_id, msg, [[
        {"text": f"✅ {neu_gesamt} importieren", "callback_data": f"heimat_ok:{uid}"},
        {"text": "❌ Verwerfen",                 "callback_data": f"heimat_no:{uid}"},
    ]])


def fetch_and_save_pending_for_url(url: str) -> dict:
    """Import für eine einzelne URL. Bekannte Gemeinde: direkt importieren.
    Neue URL: Playwright-Discovery → in heimat_gemeinden.json eintragen → importieren.
    Gibt Summary-Dict zurück (wie fetch_and_save_pending) oder {"error": "..."}."""
    gemeinden = json.loads(GEMEINDEN_FILE.read_text()) if GEMEINDEN_FILE.exists() else []

    # Bekannte Gemeinde anhand URL suchen
    treffer = next((g for g in gemeinden if g.get("url") == url), None)
    if treffer:
        return fetch_and_save_pending(gemeinden_filter=[url])

    # Neue URL: Discovery via Playwright
    _log(f"Discovery für neue URL: {url}")
    c_id = discover_c_id(url)
    if not c_id:
        return {"error": f"Keine heimat-info ID gefunden auf: {url}"}

    if any(g["c_id"] == c_id for g in gemeinden):
        # Bereits vorhanden (andere URL, gleiche c_id)
        return fetch_and_save_pending(gemeinden_filter=[next(
            g["url"] for g in gemeinden if g["c_id"] == c_id)])

    slug  = re.sub(r'https?://(www\.)?', '', url).split('/')[0]
    slug  = re.sub(r'[^a-z0-9_]', '', slug.lower().replace('-', '_').replace('.', '_'))[:20]
    name  = slug.replace('_', ' ').title()
    eintrag = {"name": name, "label": name, "verein_key": slug, "c_id": c_id, "url": url}
    gemeinden.append(eintrag)
    GEMEINDEN_FILE.write_text(json.dumps(gemeinden, ensure_ascii=False, indent=2))
    _log(f"✅ Neue Gemeinde gespeichert: {name} ({c_id[:8]}…)")

    return fetch_and_save_pending(gemeinden_filter=[url])


def main() -> None:
    secrets = load_secrets()
    if len(sys.argv) >= 3 and sys.argv[1] == "--add":
        cmd_add(sys.argv[2], secrets)
    elif len(sys.argv) >= 2 and sys.argv[1] == "--excel":
        cmd_excel(secrets)
    else:
        cmd_import(secrets)


if __name__ == "__main__":
    main()
